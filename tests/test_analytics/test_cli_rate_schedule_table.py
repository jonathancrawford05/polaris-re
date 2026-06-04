"""CLI integration tests for ``polaris rate-schedule --table`` (ADR-053).

The flat-schedule path remains covered by ``test_rate_schedule.py`` and
``polaris rate-schedule`` smoke-tests elsewhere. These tests focus on the
new ``--table`` flag, the standalone ``write_yrt_rate_table_excel``
output, the JSON serialisation helper, and the CSV-rejection guard.

All end-to-end runs solve the actual rate grid via brentq and are
``@pytest.mark.slow``; the JSON helper unit tests are fast.
"""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from polaris_re.cli import _yrt_rate_table_to_dict, app
from polaris_re.core.policy import Sex, SmokerStatus
from polaris_re.reinsurance.yrt_rate_table import YRTRateTable, YRTRateTableArray

runner = CliRunner()


# Smallest reasonable axis grid that still exercises the brentq solver
# end-to-end without making the slow tests too slow. Two ages, term 10.
_FAST_AGES = "30,40"
_FAST_TERM = "10"


@pytest.mark.slow
class TestRateScheduleTableCLI:
    """`polaris rate-schedule --table` end-to-end behaviour."""

    def test_no_table_default_runs_unchanged(self, tmp_path: Path) -> None:
        """Without ``--table`` the CSV-output path is preserved (ADR-053 backcompat)."""
        out = tmp_path / "flat.csv"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "-o",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        assert out.exists()
        # Flat schedule CSV header carries the per-cohort columns.
        first = out.read_text().splitlines()[0]
        assert "issue_age" in first
        assert "rate_per_1000" in first

    def test_table_emits_xlsx(self, tmp_path: Path) -> None:
        """`--table -o NAME.xlsx` writes a workbook with both sheets."""
        out = tmp_path / "table.xlsx"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "-o",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        assert out.exists()
        wb = load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "YRT Rate Table" in wb.sheetnames

    def test_table_csv_output_rejected(self, tmp_path: Path) -> None:
        """`--table -o NAME.csv` exits 1 with a clear error message."""
        out = tmp_path / "wrong.csv"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "-o",
                str(out),
            ],
        )
        assert result.exit_code == 1
        assert ".xlsx" in result.output
        assert not out.exists()

    def test_table_json_emits_cohort_dict(self, tmp_path: Path) -> None:
        """`--table --json PATH` writes the YRTRateTable dict shape."""
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "--json",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        payload = json.loads(out.read_text())
        # Top-level shape per ADR-053.
        assert set(payload.keys()) >= {
            "table_name",
            "min_age",
            "max_age",
            "select_period_years",
            "cohorts",
        }
        # Single demo cohort: M / U (UNKNOWN smoker, demo aggregate table).
        assert "M_U" in payload["cohorts"]
        cohort = payload["cohorts"]["M_U"]
        assert {"min_age", "max_age", "select_period", "rates"} <= set(cohort.keys())
        # Ages 30..40: n_ages = max(30,40) - min(30,40) + 1 = 11 rows.
        # Only ages 30 and 40 are solved; intermediate rows (31..39)
        # are forward/back-filled by generate_table to satisfy
        # YRTRateTableArray's contiguous-age storage contract
        # (ADR-051 documented behaviour). select_period=0 → 1 column.
        assert len(cohort["rates"]) == 11
        assert all(len(row) == 1 for row in cohort["rates"])

    def test_table_with_select_period(self, tmp_path: Path) -> None:
        """`--select-period N` produces N+1 duration columns per cohort."""
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "--select-period",
                "3",
                "--json",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        payload = json.loads(out.read_text())
        assert payload["select_period_years"] == 3
        cohort = payload["cohorts"]["M_U"]
        assert cohort["select_period"] == 3
        # Each row has select_period + 1 = 4 columns. 11 rows from the
        # contiguous age expansion (ages 30..40); see
        # test_table_json_emits_cohort_dict for the fill-in rationale.
        assert all(len(row) == 4 for row in cohort["rates"])
        # Generated table broadcasts the per-age flat rate across cols
        # (ADR-051 / ADR-053 "Out of scope") — verify the row is constant.
        for row in cohort["rates"]:
            assert max(row) == pytest.approx(min(row))


class TestYrtRateTableJsonHelper:
    """Unit tests for ``_yrt_rate_table_to_dict`` — no CLI invocation."""

    def _make_table(self) -> YRTRateTable:
        import numpy as np

        rates = np.array(
            [
                [0.50, 0.55, 0.60],
                [0.55, 0.60, 0.65],
            ],
            dtype=np.float64,
        )
        arr = YRTRateTableArray(rates=rates, min_age=40, max_age=41, select_period=2)
        return YRTRateTable.from_arrays(
            table_name="unit-test",
            arrays={(Sex.MALE, SmokerStatus.NON_SMOKER): arr},
        )

    def test_top_level_shape(self) -> None:
        """Top-level dict carries table metadata."""
        d = _yrt_rate_table_to_dict(self._make_table())
        assert d["table_name"] == "unit-test"
        assert d["min_age"] == 40
        assert d["max_age"] == 41
        assert d["select_period_years"] == 2

    def test_cohort_dict_round_trip(self) -> None:
        """Cohort entry preserves per-(age, dur) rate values."""
        d = _yrt_rate_table_to_dict(self._make_table())
        cohort = d["cohorts"]["M_NS"]
        assert cohort["min_age"] == 40
        assert cohort["max_age"] == 41
        assert cohort["select_period"] == 2
        # Rate at (age 40, dur_1) == 0.50; (age 41, ultimate) == 0.65.
        assert cohort["rates"][0][0] == pytest.approx(0.50)
        assert cohort["rates"][1][2] == pytest.approx(0.65)

    def test_dict_is_json_serialisable(self) -> None:
        """Output passes through ``json.dumps`` without errors."""
        d = _yrt_rate_table_to_dict(self._make_table())
        # Empty `default` callback so non-serialisable objects raise.
        json.dumps(d)


class TestHelperTypeGuards:
    """`_render_yrt_rate_table` and `_yrt_rate_table_to_dict` use explicit
    `raise PolarisValidationError` (not bare `assert`) so the guard holds
    under `python -O`. PR #39 P1 fix.
    """

    def test_render_rejects_non_table(self) -> None:
        from polaris_re.cli import _render_yrt_rate_table
        from polaris_re.core.exceptions import PolarisValidationError

        with pytest.raises(PolarisValidationError, match="Expected YRTRateTable"):
            _render_yrt_rate_table("not a table", target_irr=0.10)

    def test_to_dict_rejects_non_table(self) -> None:
        from polaris_re.core.exceptions import PolarisValidationError

        with pytest.raises(PolarisValidationError, match="Expected YRTRateTable"):
            _yrt_rate_table_to_dict({"not": "a table"})


class TestSolvedMaskDisclosure:
    """ADR-054 — disclosure of forward/back-filled cells in CLI / JSON output."""

    def _build_partially_solved_table(self) -> YRTRateTable:
        """Two-cohort table where one row is filled and others are solved."""
        import numpy as np

        rates = np.array([[1.0], [1.5], [2.0]], dtype=np.float64)
        # First and last rows are solved; the middle row was filled in.
        mask = np.array([[True], [False], [True]], dtype=np.bool_)
        arr = YRTRateTableArray(
            rates=rates,
            min_age=40,
            max_age=42,
            select_period=0,
            solved_mask=mask,
        )
        return YRTRateTable.from_arrays(
            table_name="partial",
            arrays={(Sex.MALE, SmokerStatus.UNKNOWN): arr},
        )

    def _build_no_mask_table(self) -> YRTRateTable:
        """CSV-loaded-style table — no provenance recorded."""
        import numpy as np

        rates = np.array([[1.0], [1.5], [2.0]], dtype=np.float64)
        arr = YRTRateTableArray(rates=rates, min_age=40, max_age=42, select_period=0)
        return YRTRateTable.from_arrays(
            table_name="no-mask",
            arrays={(Sex.MALE, SmokerStatus.UNKNOWN): arr},
        )

    def test_render_marks_filled_cells_with_asterisk(self, capsys) -> None:
        """Filled cells render with a trailing ``*``; solved cells do not."""
        from polaris_re.cli import _render_yrt_rate_table

        table = self._build_partially_solved_table()
        _render_yrt_rate_table(table, target_irr=0.10)
        out = capsys.readouterr().out
        # Solved rows render the bare 4-decimal value.
        assert "1.0000" in out
        assert "2.0000" in out
        # Filled row carries the asterisk suffix.
        assert "1.5000*" in out
        # Caption explaining the convention is printed once for the cohort.
        assert "forward/back-filled" in out

    def test_render_no_mask_is_unchanged(self, capsys) -> None:
        """CSV-loaded tables render exactly as before — no asterisks, no caption."""
        from polaris_re.cli import _render_yrt_rate_table

        table = self._build_no_mask_table()
        _render_yrt_rate_table(table, target_irr=0.10)
        out = capsys.readouterr().out
        assert "1.0000" in out
        assert "2.0000" in out
        # No asterisk on any cell.
        assert "*" not in out
        # No disclosure caption.
        assert "forward/back-filled" not in out

    def test_to_dict_includes_solved_mask_when_present(self) -> None:
        """JSON helper carries ``solved_mask`` per cohort when set."""
        d = _yrt_rate_table_to_dict(self._build_partially_solved_table())
        cohort = d["cohorts"]["M_U"]
        assert "solved_mask" in cohort
        assert cohort["solved_mask"] == [[True], [False], [True]]

    def test_to_dict_omits_solved_mask_when_absent(self) -> None:
        """JSON helper omits ``solved_mask`` when no provenance is recorded."""
        d = _yrt_rate_table_to_dict(self._build_no_mask_table())
        cohort = d["cohorts"]["M_U"]
        assert "solved_mask" not in cohort

    def test_to_dict_solved_mask_is_json_serialisable(self) -> None:
        """Ensure the mask survives ``json.dumps`` without a custom encoder."""
        d = _yrt_rate_table_to_dict(self._build_partially_solved_table())
        json.dumps(d)


@pytest.mark.slow
class TestSolvedMaskCLIIntegration:
    """End-to-end: ``rate-schedule --table`` JSON output discloses fill-in."""

    def test_sparse_ages_disclose_filled_rows_in_json(self, tmp_path: Path) -> None:
        """`--ages 30,40` writes a JSON ``solved_mask`` with True at the bookends."""
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--ages",
                _FAST_AGES,  # "30,40"
                "--term",
                _FAST_TERM,
                "--json",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        payload = json.loads(out.read_text())
        cohort = payload["cohorts"]["M_U"]
        assert "solved_mask" in cohort
        mask = cohort["solved_mask"]
        # Ages 30..40 inclusive = 11 rows; select_period=0 → 1 col.
        assert len(mask) == 11
        assert all(len(row) == 1 for row in mask)
        # Bookends solved by brentq, intermediates filled.
        assert mask[0] == [True]  # age 30
        assert mask[10] == [True]  # age 40
        for i in range(1, 10):
            assert mask[i] == [False], f"intermediate age offset {i} should be filled"


class TestSolveModeFlagValidation:
    """``--solve-mode`` flag input validation — no projection runs."""

    def test_invalid_solve_mode_value_rejected(self, tmp_path: Path) -> None:
        """Typer rejects values outside the Literal choice set."""
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--solve-mode",
                "bogus",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "--json",
                str(out),
            ],
        )
        # Typer Choice validation exits with code 2 (Click's usage error).
        assert result.exit_code != 0
        assert not out.exists()

    def test_per_duration_without_table_rejected(self, tmp_path: Path) -> None:
        """`--solve-mode per_duration` requires `--table` — error if omitted."""
        out = tmp_path / "out.csv"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--solve-mode",
                "per_duration",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "-o",
                str(out),
            ],
        )
        assert result.exit_code == 1
        assert "--table" in result.output
        assert not out.exists()

    def test_flat_solve_mode_without_table_runs_unchanged(self, tmp_path: Path) -> None:
        """The default `flat` mode is the existing no-op when `--table` is unset."""
        out = tmp_path / "flat.csv"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--solve-mode",
                "flat",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "-o",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        assert out.exists()


@pytest.mark.slow
class TestSolveModePerDurationCLI:
    """End-to-end: ``rate-schedule --table --solve-mode per_duration``."""

    def test_per_duration_table_name_carries_suffix(self, tmp_path: Path) -> None:
        """`table_name` ends with `_per_duration` so reviewers can identify the mode."""
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--solve-mode",
                "per_duration",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "--select-period",
                "2",
                "--json",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        payload = json.loads(out.read_text())
        assert payload["table_name"].endswith("_per_duration")

    def test_flat_table_name_carries_flat_suffix(self, tmp_path: Path) -> None:
        """Default `flat` mode tags `table_name` with `_flat` (ADR-063)."""
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "--json",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        payload = json.loads(out.read_text())
        assert payload["table_name"].endswith("_flat")

    def test_per_duration_produces_per_cell_rates(self, tmp_path: Path) -> None:
        """`--solve-mode per_duration` with select-period > 0 yields non-uniform rows.

        Under `--solve-mode flat` (default) every column of a given age row
        is the same broadcast rate. Under `per_duration` the solver runs
        independently per (age, duration) cell, so at least one column
        should differ from column 0 on the demo synthetic table.
        """
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--solve-mode",
                "per_duration",
                "--ages",
                _FAST_AGES,
                "--term",
                _FAST_TERM,
                "--select-period",
                "3",
                "--json",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        payload = json.loads(out.read_text())
        cohort = payload["cohorts"]["M_U"]
        assert cohort["select_period"] == 3
        # At least one solved row should have a non-uniform column profile.
        # (Filled rows are forward/back-filled and may be uniform; the
        # bookend rows for ages 30 and 40 are solved.)
        rates = cohort["rates"]
        mask = cohort["solved_mask"]
        solved_rows = [r for r, m in zip(rates, mask, strict=True) if all(m)]
        assert solved_rows, "expected at least one fully solved row"
        non_uniform = [r for r in solved_rows if max(r) != min(r)]
        assert non_uniform, (
            "per_duration mode should produce at least one non-uniform "
            f"solved row, got rows={solved_rows}"
        )

    def test_per_duration_solved_mask_is_per_cell(self, tmp_path: Path) -> None:
        """Under `per_duration` the solved_mask is genuinely 2-D per-cell.

        For sparse age input each requested-age row is solved cell-by-cell
        (every column True), while filled rows show all-False — matching
        the analytics-layer contract in
        `test_per_duration_sparse_ages_mark_only_solved_cells`.
        """
        out = tmp_path / "table.json"
        result = runner.invoke(
            app,
            [
                "rate-schedule",
                "--table",
                "--solve-mode",
                "per_duration",
                "--ages",
                _FAST_AGES,  # "30,40"
                "--term",
                _FAST_TERM,
                "--select-period",
                "2",
                "--json",
                str(out),
            ],
        )
        assert result.exit_code == 0, result.output
        payload = json.loads(out.read_text())
        mask = payload["cohorts"]["M_U"]["solved_mask"]
        # Ages 30..40 = 11 rows; select_period=2 → 3 cols.
        assert len(mask) == 11
        assert all(len(row) == 3 for row in mask)
        # Bookends (rows 0 and 10) — every column solved.
        assert mask[0] == [True, True, True]
        assert mask[10] == [True, True, True]
        # Filled rows — every column False.
        for i in range(1, 10):
            assert mask[i] == [False, False, False]
