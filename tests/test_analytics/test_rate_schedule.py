"""Tests for YRT Rate Schedule Generator."""

from datetime import date
from pathlib import Path

import numpy as np
import pytest

from polaris_re.analytics.rate_schedule import YRTRateSchedule
from polaris_re.assumptions.assumption_set import AssumptionSet
from polaris_re.assumptions.lapse import LapseAssumption
from polaris_re.assumptions.mortality import MortalityTable, MortalityTableSource
from polaris_re.core.policy import Sex, SmokerStatus
from polaris_re.core.projection import ProjectionConfig
from polaris_re.utils.table_io import load_mortality_csv

FIXTURES = Path(__file__).parent.parent / "fixtures"


@pytest.fixture()
def assumptions():
    """Create assumptions for rate schedule tests."""
    table_array = load_mortality_csv(
        FIXTURES / "synthetic_select_ultimate.csv",
        select_period=3,
        min_age=18,
        max_age=60,
    )
    mortality = MortalityTable.from_table_array(
        source=MortalityTableSource.SOA_VBT_2015,
        table_name="Synthetic",
        table_array=table_array,
        sex=Sex.MALE,
        smoker_status=SmokerStatus.UNKNOWN,
    )
    lapse = LapseAssumption.from_duration_table(
        {1: 0.10, 2: 0.08, 3: 0.06, 4: 0.05, 5: 0.04, "ultimate": 0.03}
    )
    return AssumptionSet(mortality=mortality, lapse=lapse, version="test-schedule")


@pytest.fixture()
def config():
    """Projection config for rate schedule tests."""
    return ProjectionConfig(
        projection_horizon_years=20,
        discount_rate=0.05,
        valuation_date=date(2025, 1, 1),
    )


class TestYRTRateSchedule:
    """Tests for rate schedule generation."""

    def test_basic_generation(self, assumptions, config):
        """Generates a rate schedule for a few ages."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config, target_irr=0.10)
        df = scheduler.generate(
            ages=[35, 45],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
        )
        assert len(df) == 2
        assert "rate_per_1000" in df.columns
        assert "issue_age" in df.columns

    def test_rates_are_positive(self, assumptions, config):
        """Solved rates should be positive."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config, target_irr=0.10)
        df = scheduler.generate(
            ages=[35, 40, 45],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
        )
        rates = df["rate_per_1000"].to_numpy()
        valid_rates = rates[~np.isnan(rates)]
        assert len(valid_rates) > 0
        assert np.all(valid_rates > 0)

    def test_rates_increase_with_age(self, assumptions, config):
        """Rates should generally increase with age (higher mortality = higher rate)."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config, target_irr=0.10)
        df = scheduler.generate(
            ages=[30, 40, 50],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
        )
        rates = df["rate_per_1000"].to_numpy()
        valid = rates[~np.isnan(rates)]
        if len(valid) >= 2:
            # At minimum, oldest should have higher rate than youngest
            assert valid[-1] > valid[0]

    def test_output_columns(self, assumptions, config):
        """Output DataFrame has expected columns."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        df = scheduler.generate(ages=[40], sexes=[Sex.MALE], smoker_statuses=[SmokerStatus.UNKNOWN])
        expected = {"issue_age", "sex", "smoker_status", "policy_term", "rate_per_1000", "irr"}
        assert set(df.columns) == expected

    def test_grid_size(self, assumptions, config):
        """Output has correct number of rows for the grid."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        df = scheduler.generate(
            ages=[30, 40, 50],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
        )
        assert len(df) == 3  # 3 ages x 1 sex x 1 smoker

    def test_different_target_irr(self, assumptions, config):
        """Different target IRR produces different rates."""
        sched_low = YRTRateSchedule(assumptions=assumptions, config=config, target_irr=0.08)
        sched_high = YRTRateSchedule(assumptions=assumptions, config=config, target_irr=0.15)

        df_low = sched_low.generate(
            ages=[40], sexes=[Sex.MALE], smoker_statuses=[SmokerStatus.UNKNOWN]
        )
        df_high = sched_high.generate(
            ages=[40], sexes=[Sex.MALE], smoker_statuses=[SmokerStatus.UNKNOWN]
        )

        rate_low = df_low["rate_per_1000"][0]
        rate_high = df_high["rate_per_1000"][0]

        if not (np.isnan(rate_low) or np.isnan(rate_high)):
            # Rates should differ for different target IRRs
            assert rate_low != rate_high
            # Higher hurdle rate discounts future claim outflows more, so
            # the reinsurer can accept a lower rate and still hit IRR
            assert rate_high < rate_low


class TestGenerateTable:
    """Tests for `YRTRateSchedule.generate_table()` (Slice 2 of YRT rate table)."""

    def test_generate_table_returns_yrt_rate_table(self, assumptions, config):
        """generate_table() returns a populated YRTRateTable."""
        from polaris_re.reinsurance import YRTRateTable

        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[35, 40, 45],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=0,
        )
        assert isinstance(table, YRTRateTable)
        assert table.min_age == 35
        assert table.max_age == 45
        assert table.select_period_years == 0
        # The (MALE, UNKNOWN) cohort must be loaded.
        assert "M_U" in table.arrays

    def test_generate_table_round_trips_through_treaty(self, assumptions, config):
        """A generated table fed back into YRTTreaty.apply() produces a
        non-empty ceded premium series with no errors. This is the closed-
        loop sanity check for the Slice 2 contract."""
        from polaris_re.core.inforce import InforceBlock
        from polaris_re.core.policy import Policy, ProductType
        from polaris_re.products.term_life import TermLife
        from polaris_re.reinsurance import YRTTreaty

        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=0,
        )

        policy = Policy(
            policy_id="ROUND_TRIP",
            issue_age=40,
            attained_age=40,
            sex=Sex.MALE,
            smoker_status=SmokerStatus.UNKNOWN,
            underwriting_class="STANDARD",
            face_amount=1_000_000.0,
            annual_premium=12_000.0,
            product_type=ProductType.TERM,
            policy_term=20,
            duration_inforce=0,
            reinsurance_cession_pct=1.0,
            issue_date=date(2025, 1, 1),
            valuation_date=date(2025, 1, 1),
        )
        block = InforceBlock(policies=[policy])
        gross = TermLife(block, assumptions, config).project(seriatim=True)

        treaty = YRTTreaty(
            cession_pct=1.0,
            total_face_amount=1_000_000.0,
            yrt_rate_table=table,
        )
        net, ceded = treaty.apply(gross, inforce=block)
        # No NaNs / Infs allowed in either output.
        assert np.all(np.isfinite(ceded.gross_premiums))
        assert np.all(np.isfinite(net.net_cash_flow))
        # Some ceded premium must be collected.
        assert ceded.gross_premiums.sum() > 0


class TestGenerateTableSolvedMask:
    """ADR-054 — ``generate_table()`` records per-cell solver provenance.

    Sparse age inputs (e.g. ``ages=[30, 40]``) produce a table whose
    ``YRTRateTableArray`` storage is contiguous from min to max age. The
    ``solved_mask`` distinguishes brentq-solved rows (True) from rows
    that were forward/back-filled to satisfy the contiguous-storage
    contract (False). Renderers consume this to disclose interpolation;
    consumption (``YRTTreaty.apply``) does not branch on it.
    """

    def test_dense_grid_is_fully_solved(self, assumptions, config):
        """Every requested age is solved, so the mask is all True."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[35, 36, 37],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=0,
        )
        arr = table.arrays["M_U"]
        assert arr.solved_mask is not None
        assert arr.is_fully_solved
        assert arr.solved_mask.shape == (3, 1)
        assert bool(arr.solved_mask.all())

    def test_sparse_grid_marks_filled_rows(self, assumptions, config):
        """Unrequested intermediate ages are marked False in the mask."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[35, 40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=0,
        )
        arr = table.arrays["M_U"]
        # Storage is contiguous: ages 35..40 inclusive (6 rows). Only the
        # bookend rows were solved by brentq.
        assert arr.solved_mask is not None
        assert arr.solved_mask.shape == (6, 1)
        assert bool(arr.solved_mask[0, 0])  # age 35 — solved
        assert bool(arr.solved_mask[5, 0])  # age 40 — solved
        # Intermediate ages 36..39 were forward/back-filled.
        for age_offset in (1, 2, 3, 4):
            assert not bool(arr.solved_mask[age_offset, 0])
        assert not arr.is_fully_solved

    def test_mask_broadcasts_across_select_columns(self, assumptions, config):
        """All select-period columns share the same per-row solved status.

        Because ``generate_table`` broadcasts the per-age flat rate across
        every duration column (ADR-051 / ADR-053 "Out of scope"), the
        per-cell solved flag is also row-uniform: every column in a solved
        row is True; every column in a filled row is False.
        """
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[35, 40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=3,
        )
        arr = table.arrays["M_U"]
        assert arr.solved_mask is not None
        assert arr.solved_mask.shape == (6, 4)
        for age_offset in range(arr.solved_mask.shape[0]):
            row = arr.solved_mask[age_offset, :]
            assert bool(row.all()) or not bool(row.any()), (
                f"row {age_offset} mask is mixed True/False; "
                "broadcast contract requires uniform-by-row"
            )


class TestGenerateTablePerDuration:
    """ADR-063 — ``generate_table(solve_mode="per_duration")`` solves each
    (age, duration) cell independently.

    The default ``solve_mode="flat"`` mode (covered by ``TestGenerateTable``
    above) solves one flat rate per (age, sex, smoker) and broadcasts it
    across every duration column — the row-uniform contract enforced by
    ``test_mask_broadcasts_across_select_columns``. The ``"per_duration"``
    mode solves a separate rate per cell by projecting a synthetic policy
    that has been inforce for ``d`` years at the row's issue age, lighting
    up ``solved_mask`` as a genuinely 2-D per-cell map.
    """

    def test_per_duration_yields_distinct_rates_across_columns(self, assumptions, config):
        """Per-duration mode produces different rates per column (not a broadcast)."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=3,
            solve_mode="per_duration",
        )
        arr = table.arrays["M_U"]
        assert arr.rates.shape == (1, 4)
        row = arr.rates[0, :]
        # Per-duration mode: at least one column should differ from column 0
        # (flat mode would have all four columns identical).
        assert not np.allclose(row, row[0]), (
            f"Per-duration rates should not be row-uniform, got {row}"
        )

    def test_per_duration_rates_increase_within_select_period(self, assumptions, config):
        """Rates rise across the select period as underwriting durability wears off.

        The fixture's mortality is select-distinct and ascending with
        duration (dur_1 < dur_2 < dur_3 < ultimate), so the solved YRT
        rate at duration d should also increase with d.
        """
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=3,
            solve_mode="per_duration",
        )
        row = table.arrays["M_U"].rates[0, :]
        for d in range(len(row) - 1):
            assert row[d + 1] >= row[d], (
                f"YRT rate at duration {d + 1} ({row[d + 1]:.4f}) should be "
                f">= rate at duration {d} ({row[d]:.4f}) given select-rising mortality"
            )
        # And the rate at the ultimate column should strictly exceed the rate
        # at duration 0 — the fixture's ultimate rate is markedly higher than
        # any select cell, so the IRR solver must reflect that.
        assert row[-1] > row[0]

    def test_per_duration_dense_grid_is_fully_solved(self, assumptions, config):
        """A dense grid in per-duration mode yields an all-True 2-D mask."""
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[35, 36, 37],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=2,
            solve_mode="per_duration",
        )
        arr = table.arrays["M_U"]
        assert arr.solved_mask is not None
        assert arr.solved_mask.shape == (3, 3)
        assert bool(arr.solved_mask.all())
        assert arr.is_fully_solved

    def test_per_duration_sparse_ages_mark_only_solved_cells(self, assumptions, config):
        """Sparse age input: only the requested-age rows have True cells.

        Forward/back-filled age rows must show False across every column,
        and the requested rows must show True across every column (each
        column was solved independently for that age).
        """
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[35, 40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=2,
            solve_mode="per_duration",
        )
        arr = table.arrays["M_U"]
        assert arr.solved_mask is not None
        assert arr.solved_mask.shape == (6, 3)
        # Requested ages: rows 0 (age 35) and 5 (age 40) — every column True.
        assert bool(arr.solved_mask[0, :].all())
        assert bool(arr.solved_mask[5, :].all())
        # Filled ages: rows 1..4 (ages 36..39) — every column False.
        for age_offset in (1, 2, 3, 4):
            assert not bool(arr.solved_mask[age_offset, :].any())
        assert not arr.is_fully_solved

    def test_per_duration_select_period_zero_matches_flat(self, assumptions, config):
        """At select_period_years=0 both modes reduce to the same single-column solve.

        The flat-mode solver runs once per (age, sex, smoker) and broadcasts.
        At a one-column table there is nothing to broadcast, and the
        per-duration solve at d=0 is the same underlying optimisation, so
        the produced rates must agree to within solver tolerance.
        """
        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        flat = scheduler.generate_table(
            ages=[35, 40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=0,
            solve_mode="flat",
        )
        per_dur = scheduler.generate_table(
            ages=[35, 40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=0,
            solve_mode="per_duration",
        )
        np.testing.assert_allclose(
            per_dur.arrays["M_U"].rates,
            flat.arrays["M_U"].rates,
            rtol=1e-3,
        )

    def test_per_duration_round_trips_through_treaty(self, assumptions, config):
        """A per-duration table feeds back into YRTTreaty.apply without errors.

        Closed-loop check: synthetic table consumed by YRTTreaty.apply
        produces finite ceded cash flows and a non-zero ceded premium.
        """
        from polaris_re.core.inforce import InforceBlock
        from polaris_re.core.policy import Policy, ProductType
        from polaris_re.products.term_life import TermLife
        from polaris_re.reinsurance import YRTTreaty

        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        table = scheduler.generate_table(
            ages=[40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
            policy_term=20,
            select_period_years=2,
            solve_mode="per_duration",
        )

        policy = Policy(
            policy_id="ROUND_TRIP_PD",
            issue_age=40,
            attained_age=40,
            sex=Sex.MALE,
            smoker_status=SmokerStatus.UNKNOWN,
            underwriting_class="STANDARD",
            face_amount=1_000_000.0,
            annual_premium=12_000.0,
            product_type=ProductType.TERM,
            policy_term=20,
            duration_inforce=0,
            reinsurance_cession_pct=1.0,
            issue_date=date(2025, 1, 1),
            valuation_date=date(2025, 1, 1),
        )
        block = InforceBlock(policies=[policy])
        gross = TermLife(block, assumptions, config).project(seriatim=True)

        treaty = YRTTreaty(
            cession_pct=1.0,
            total_face_amount=1_000_000.0,
            yrt_rate_table=table,
        )
        net, ceded = treaty.apply(gross, inforce=block)
        assert np.all(np.isfinite(ceded.gross_premiums))
        assert np.all(np.isfinite(net.net_cash_flow))
        assert ceded.gross_premiums.sum() > 0

    def test_invalid_solve_mode_raises(self, assumptions, config):
        """Unknown solve_mode is rejected up-front (no silent fallback)."""
        from polaris_re.core.exceptions import PolarisValidationError

        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        with pytest.raises(PolarisValidationError):
            scheduler.generate_table(
                ages=[40],
                sexes=[Sex.MALE],
                smoker_statuses=[SmokerStatus.UNKNOWN],
                policy_term=20,
                select_period_years=2,
                solve_mode="bogus",  # type: ignore[arg-type]
            )


class TestExcelOutput:
    """Tests for Excel rate schedule export."""

    def test_excel_output(self, assumptions, config, tmp_path):
        """Excel file is created with correct structure."""
        from polaris_re.utils.excel_output import write_rate_schedule_excel

        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        df = scheduler.generate(
            ages=[35, 45],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
        )

        excel_path = tmp_path / "rates.xlsx"
        write_rate_schedule_excel(df, excel_path)

        assert excel_path.exists()
        assert excel_path.stat().st_size > 0

    def test_excel_has_sheets(self, assumptions, config, tmp_path):
        """Excel workbook has correct sheet names."""
        from openpyxl import load_workbook

        from polaris_re.utils.excel_output import write_rate_schedule_excel

        scheduler = YRTRateSchedule(assumptions=assumptions, config=config)
        df = scheduler.generate(
            ages=[40],
            sexes=[Sex.MALE],
            smoker_statuses=[SmokerStatus.UNKNOWN],
        )

        excel_path = tmp_path / "rates.xlsx"
        write_rate_schedule_excel(df, excel_path)

        wb = load_workbook(excel_path)
        assert "Summary" in wb.sheetnames
        assert "M_U" in wb.sheetnames
