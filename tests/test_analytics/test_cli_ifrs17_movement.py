"""
CLI IFRS 17 movement-table surfacing tests (IFRS 17 epic, Slice 3c — ADR-097).

``polaris price --ifrs17-movement`` emits the IFRS 17 analysis-of-change
(movement) table per product cohort: the cohort's policies are grouped into
annual issue-year cohorts, each measured BBA at the config discount rate
(locked-in) and rolled forward opening → new business → interest accretion →
release → closing for BEL / RA / CSM. The table is added to the JSON output, the
Rich terminal output, and — with ``--excel-out`` — the workbook.

The flag is opt-in; runs without it carry no ``ifrs17_movement`` key (the
golden-baseline guarantee). The JSON shape mirrors the REST
``IFRS17MovementResponse`` (ADR-095) so the CLI and API agree.
"""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from polaris_re.cli import app

runner = CliRunner()

REPO_ROOT = Path(__file__).parent.parent.parent
GOLDEN_DIR = REPO_ROOT / "data" / "qa"
GOLDEN_CSV = GOLDEN_DIR / "golden_inforce.csv"
GOLDEN_CONFIG = GOLDEN_DIR / "golden_config_flat.json"

# golden_config_flat uses discount_rate 0.06; that is the locked-in rate the CLI
# applies to every issue-year cohort (a per-year override is a promoted
# follow-up).
GOLDEN_DISCOUNT_RATE = 0.06

_MOVEMENT_KEYS = {"months_per_period", "n_cohorts", "max_footing_error", "aggregate", "cohorts"}
_COMPONENT_KEYS = {"opening", "new_business", "interest_accretion", "release", "closing"}

# A single-product (TERM), single-issue-year inforce block → exactly one IFRS 17
# issue-year cohort, so the movement table is mirrored at the JSON top level.
# Written to tmp_path so the test is hermetic: it does not depend on the demo
# fixtures, which are not shipped in the runtime Docker image that runs the suite.
_SINGLE_COHORT_CSV = (
    "policy_id,issue_age,attained_age,sex,smoker_status,underwriting_class,"
    "face_amount,annual_premium,product_type,policy_term,duration_inforce,"
    "reinsurance_cession_pct,issue_date,valuation_date\n"
    "SC-001,30,35,M,NS,PREFERRED,500000.00,300.00,TERM,20,60,,2021-04-01,2026-04-01\n"
    "SC-002,45,50,F,NS,STANDARD,1000000.00,1200.00,TERM,20,60,,2021-04-01,2026-04-01\n"
)


def _single_cohort_csv(tmp_path: Path) -> Path:
    """Write a single-product, single-issue-year inforce CSV under ``tmp_path``."""
    csv = tmp_path / "single_cohort_inforce.csv"
    csv.write_text(_SINGLE_COHORT_CSV)
    return csv


def _run(csv: Path, config: Path, tmp_path: Path, *extra: str) -> dict:
    out = tmp_path / "result.json"
    result = runner.invoke(
        app,
        [
            "price",
            "--config",
            str(config),
            "--inforce",
            str(csv),
            "--output",
            str(out),
            *extra,
        ],
    )
    assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
    return json.loads(out.read_text())


class TestCLIIFRS17MovementBackwardCompat:
    """Without the flag the output is unchanged — no ``ifrs17_movement`` key."""

    def test_no_flag_no_cohort_key(self, tmp_path: Path) -> None:
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path)
        assert payload["cohorts"]
        assert all("ifrs17_movement" not in c for c in payload["cohorts"])

    def test_no_flag_no_top_level_key(self, tmp_path: Path) -> None:
        payload = _run(_single_cohort_csv(tmp_path), GOLDEN_CONFIG, tmp_path)
        assert "ifrs17_movement" not in payload


class TestCLIIFRS17MovementJSON:
    """``--ifrs17-movement`` adds the movement block per cohort and at top level."""

    def test_each_cohort_has_movement_block(self, tmp_path: Path) -> None:
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        assert payload["cohorts"]
        for cohort in payload["cohorts"]:
            assert "ifrs17_movement" in cohort
            assert set(cohort["ifrs17_movement"]) == _MOVEMENT_KEYS

    def test_single_cohort_top_level_mirror(self, tmp_path: Path) -> None:
        # A single-product, single-issue-year block → one IFRS 17 cohort,
        # mirrored at the top level for the common single-cohort case.
        payload = _run(_single_cohort_csv(tmp_path), GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        assert set(payload["ifrs17_movement"]) == _MOVEMENT_KEYS
        assert payload["ifrs17_movement"]["n_cohorts"] == 1

    def test_table_foots(self, tmp_path: Path) -> None:
        """Headline disclosure property: opening + Σ movements == closing."""
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        for cohort in payload["cohorts"]:
            assert cohort["ifrs17_movement"]["max_footing_error"] < 1e-6

    def test_cohorts_grouped_and_ordered_by_issue_year(self, tmp_path: Path) -> None:
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        by_product = {c["product_type"]: c["ifrs17_movement"] for c in payload["cohorts"]}
        # golden_inforce: TERM issued 2021 & 2026; WHOLE_LIFE issued 2016, 2021 & 2026.
        term_years = [t["issue_year"] for t in by_product["TERM"]["cohorts"]]
        wl_years = [t["issue_year"] for t in by_product["WHOLE_LIFE"]["cohorts"]]
        assert term_years == [2021, 2026]
        assert wl_years == [2016, 2021, 2026]

    def test_aggregate_has_null_cohort_metadata(self, tmp_path: Path) -> None:
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        aggregate = payload["cohorts"][0]["ifrs17_movement"]["aggregate"]
        assert aggregate["issue_year"] is None
        assert aggregate["locked_in_rate"] is None

    def test_locked_in_rate_is_config_discount_rate(self, tmp_path: Path) -> None:
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        for cohort in payload["cohorts"]:
            for table in cohort["ifrs17_movement"]["cohorts"]:
                assert table["locked_in_rate"] == pytest.approx(GOLDEN_DISCOUNT_RATE)

    def test_rows_carry_all_components(self, tmp_path: Path) -> None:
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        row = payload["cohorts"][0]["ifrs17_movement"]["aggregate"]["rows"][0]
        assert {"bel", "ra", "csm", "total"}.issubset(row)
        assert _COMPONENT_KEYS.issubset(row["bel"])

    def test_annual_reporting_periods_default(self, tmp_path: Path) -> None:
        payload = _run(GOLDEN_CSV, GOLDEN_CONFIG, tmp_path, "--ifrs17-movement")
        movement = payload["cohorts"][0]["ifrs17_movement"]
        assert movement["months_per_period"] == 12
        # 20-year projection horizon → 20 annual reporting periods.
        assert movement["aggregate"]["n_periods"] == 20

    def test_months_per_period_override(self, tmp_path: Path) -> None:
        payload = _run(
            GOLDEN_CSV,
            GOLDEN_CONFIG,
            tmp_path,
            "--ifrs17-movement",
            "--ifrs17-months-per-period",
            "6",
        )
        movement = payload["cohorts"][0]["ifrs17_movement"]
        assert movement["months_per_period"] == 6
        # 240 months / 6 → 40 reporting periods.
        assert movement["aggregate"]["n_periods"] == 40
        # Still foots after re-aggregation.
        assert movement["max_footing_error"] < 1e-6


class TestCLIIFRS17MovementValidation:
    """Out-of-range flag values fail fast with a clean CLI error."""

    def test_ra_factor_out_of_range_rejected(self, tmp_path: Path) -> None:
        out = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIG),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(out),
                "--ifrs17-movement",
                "--ifrs17-ra-factor",
                "0.9",
            ],
        )
        assert result.exit_code != 0

    def test_months_per_period_zero_rejected(self, tmp_path: Path) -> None:
        out = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIG),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(out),
                "--ifrs17-movement",
                "--ifrs17-months-per-period",
                "0",
            ],
        )
        assert result.exit_code != 0


class TestCLIIFRS17MovementExcel:
    """``--excel-out`` with the flag appends the 'IFRS 17 Movement' sheet."""

    def test_excel_sheet_appended_per_cohort(self, tmp_path: Path) -> None:
        out_json = tmp_path / "result.json"
        excel_out = tmp_path / "deal.xlsx"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIG),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(out_json),
                "--excel-out",
                str(excel_out),
                "--ifrs17-movement",
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        # One workbook per product cohort (cohort id appended to the stem).
        workbooks = sorted(tmp_path.glob("deal-*.xlsx"))
        assert workbooks, "no per-cohort workbooks written"
        for wb_path in workbooks:
            wb = load_workbook(wb_path)
            assert "IFRS 17 Movement" in wb.sheetnames

    def test_excel_sheet_absent_without_flag(self, tmp_path: Path) -> None:
        out_json = tmp_path / "result.json"
        excel_out = tmp_path / "deal.xlsx"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIG),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(out_json),
                "--excel-out",
                str(excel_out),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        for wb_path in sorted(tmp_path.glob("deal-*.xlsx")):
            wb = load_workbook(wb_path)
            assert "IFRS 17 Movement" not in wb.sheetnames
