"""CLI end-to-end tests against golden inputs.

Uses typer.CliRunner to invoke polaris price/scenario/uq commands
with the golden inforce CSV and config files, asserting on exit code
and output structure.
"""

import json

import pytest
from openpyxl import load_workbook
from typer.testing import CliRunner

from polaris_re.cli import app

from .conftest import GOLDEN_CONFIGS_DIR, GOLDEN_CSV, requires_soa_tables

runner = CliRunner()


class TestCLIGoldenSmoke:
    """Smoke tests: CLI commands run to completion on golden inputs."""

    def test_price_flat_mortality(self, tmp_path):
        """polaris price runs on golden CSV with flat mortality."""
        output = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(output),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        payload = json.loads(output.read_text())
        assert "cohorts" in payload
        assert payload["summary"]["n_cohorts"] == 2

    @requires_soa_tables
    def test_price_yrt_soa(self, tmp_path):
        """polaris price runs with SOA VBT 2015 tables."""
        output = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_yrt.json"),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(output),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        payload = json.loads(output.read_text())
        assert payload["summary"]["n_cohorts"] == 2
        # Both cohorts should have cedant profit test results
        for cohort in payload["cohorts"]:
            assert "pv_profits" in cohort["cedant"]

    @requires_soa_tables
    def test_price_coinsurance(self, tmp_path):
        """polaris price runs with coinsurance treaty."""
        output = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_coins.json"),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(output),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"

    @requires_soa_tables
    def test_price_policy_cession(self, tmp_path):
        """polaris price runs with policy-level cession overrides."""
        output = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_policy_cession.json"),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(output),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"

    def test_scenario_rejects_mixed_block(self, tmp_path):
        """polaris scenario exits non-zero on mixed product block."""
        result = runner.invoke(
            app,
            [
                "scenario",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(GOLDEN_CSV),
            ],
        )
        assert result.exit_code != 0

    def test_validate_golden_csv(self):
        """polaris validate accepts the golden CSV."""
        result = runner.invoke(
            app,
            [
                "validate",
                str(GOLDEN_CSV),
            ],
        )
        assert result.exit_code == 0


class TestCLIRatedBlockOutput:
    """Slice 3: CLI emits rating composition on `polaris price` JSON output."""

    def test_all_standard_block_reports_zero_rated(self, tmp_path):
        """Golden CSV has no substandard lives → rated_block shows 0 rated."""
        output = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(output),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        payload = json.loads(output.read_text())
        assert "rated_block" in payload
        rb = payload["rated_block"]
        assert rb["n_rated"] == 0
        assert rb["pct_rated_by_count"] == 0.0
        assert rb["face_weighted_mean_multiplier"] == 1.0

    def test_rated_csv_surfaces_rating_composition(self, tmp_path):
        """A CSV with substandard lives populates rated_block and renders the table."""
        # Build a rated CSV from the golden block plus two substandard rows.
        import polars as pl

        golden = pl.read_csv(GOLDEN_CSV, try_parse_dates=False)
        # Add the new columns with defaults on existing rows
        golden = golden.with_columns(
            [
                pl.lit(1.0, dtype=pl.Float64).alias("mortality_multiplier"),
                pl.lit(0.0, dtype=pl.Float64).alias("flat_extra_per_1000"),
            ]
        )
        rated_rows = [
            {
                "policy_id": "RATED-TBL2",
                "issue_age": 35,
                "attained_age": 40,
                "sex": "M",
                "smoker_status": "NS",
                "underwriting_class": "SUBSTANDARD",
                "face_amount": 500_000.0,
                "annual_premium": 1200.0,
                "product_type": "TERM",
                "policy_term": 20,
                "duration_inforce": 60,
                "reinsurance_cession_pct": None,
                "mortality_multiplier": 2.0,
                "flat_extra_per_1000": 0.0,
                "issue_date": "2021-04-01",
                "valuation_date": "2026-04-01",
            },
            {
                "policy_id": "RATED-FE5",
                "issue_age": 45,
                "attained_age": 50,
                "sex": "F",
                "smoker_status": "NS",
                "underwriting_class": "SUBSTANDARD",
                "face_amount": 500_000.0,
                "annual_premium": 2000.0,
                "product_type": "TERM",
                "policy_term": 20,
                "duration_inforce": 60,
                "reinsurance_cession_pct": None,
                "mortality_multiplier": 1.0,
                "flat_extra_per_1000": 5.0,
                "issue_date": "2021-04-01",
                "valuation_date": "2026-04-01",
            },
        ]
        rated_df = pl.concat(
            [golden, pl.DataFrame(rated_rows, schema=golden.schema)],
            how="vertical",
        )
        rated_csv = tmp_path / "rated_inforce.csv"
        rated_df.write_csv(rated_csv)

        output = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(rated_csv),
                "--output",
                str(output),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        payload = json.loads(output.read_text())
        rb = payload["rated_block"]
        assert rb["n_rated"] == 2
        assert rb["max_multiplier"] == 2.0
        assert rb["max_flat_extra_per_1000"] == 5.0
        # % by count = 2 / (6 + 2 added) — 6 golden + 2 rated = 8 total
        assert rb["pct_rated_by_count"] > 0.0


class TestCLIExcelOut:
    """Slice 2 of ADR-045/046: ``polaris price --excel-out`` wiring."""

    def test_price_excel_out_single_cohort(self, tmp_path):
        """`polaris price --excel-out` writes a workbook whose Summary
        IRR cell matches the JSON ``cedant.irr`` field."""
        # Build a TERM-only inforce CSV so the run produces a single cohort.
        import polars as pl

        df = pl.read_csv(GOLDEN_CSV, try_parse_dates=False)
        term_only = df.filter(pl.col("product_type") == "TERM")
        term_csv = tmp_path / "term_only_inforce.csv"
        term_only.write_csv(term_csv)

        json_out = tmp_path / "result.json"
        xlsx_out = tmp_path / "deal.xlsx"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(term_csv),
                "--output",
                str(json_out),
                "--excel-out",
                str(xlsx_out),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        assert xlsx_out.exists(), "single-cohort run must write exactly the supplied path"

        payload = json.loads(json_out.read_text())
        assert payload["summary"]["n_cohorts"] == 1
        expected_irr = payload["cohorts"][0]["cedant"]["irr"]

        wb = load_workbook(xlsx_out)
        assert set(wb.sheetnames) >= {"Summary", "Cash Flows", "Assumptions"}
        assert "Sensitivity" not in wb.sheetnames  # scenarios not populated by `polaris price`

        ws = wb["Summary"]
        # Metric column laid out by _write_summary_sheet: header at row 3,
        # metrics start at row 4 in the fixed order from _SUMMARY_METRICS.
        # IRR is the 5th metric → row 8, column 2 (Cedant).
        assert ws.cell(row=3, column=1).value == "Metric"
        assert ws.cell(row=3, column=2).value == "Cedant (NET)"
        assert ws.cell(row=8, column=1).value == "IRR"
        cell_irr = ws.cell(row=8, column=2).value
        if expected_irr is None:
            assert cell_irr == "N/A"
        else:
            assert cell_irr == pytest.approx(expected_irr, rel=1e-9)

        cf = wb["Cash Flows"]
        # Row 1 = header; data rows should equal projection_years. Flat
        # config sets projection_years=20.
        data_rows = [row for row in cf.iter_rows(min_row=2, values_only=True) if row[0] is not None]
        assert len(data_rows) == 20
        assert cf.cell(row=1, column=1).value == "Year"
        assert cf.cell(row=1, column=7).value == "Net Cash Flow"

    def test_price_excel_out_mixed_cohort_writes_one_file_per_cohort(self, tmp_path):
        """Mixed-cohort blocks emit one workbook per cohort, each with
        the cohort id appended to the stem."""
        json_out = tmp_path / "result.json"
        xlsx_out = tmp_path / "deal.xlsx"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(json_out),
                "--excel-out",
                str(xlsx_out),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        payload = json.loads(json_out.read_text())
        cohort_ids = [c["product_type"] for c in payload["cohorts"]]
        assert set(cohort_ids) == {"TERM", "WHOLE_LIFE"}

        # Base path should NOT exist — single-file mode is reserved for
        # single-cohort runs.
        assert not xlsx_out.exists()
        for cid in cohort_ids:
            per_cohort = tmp_path / f"deal-{cid}.xlsx"
            assert per_cohort.exists(), f"missing per-cohort workbook: {per_cohort.name}"
            wb = load_workbook(per_cohort)
            assert "Summary" in wb.sheetnames
            assert "Cash Flows" in wb.sheetnames
            assert "Assumptions" in wb.sheetnames

    def test_price_without_excel_out_writes_no_workbook(self, tmp_path):
        """No `--excel-out` flag → JSON-only behaviour (no regression)."""
        json_out = tmp_path / "result.json"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(GOLDEN_CSV),
                "--output",
                str(json_out),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        assert json_out.exists()
        # No .xlsx anywhere under the tmp tree.
        assert not list(tmp_path.glob("*.xlsx"))

    def test_price_excel_out_assumptions_sheet_reflects_config(self, tmp_path):
        """Assumptions sheet carries the deal's hurdle rate and treaty info."""
        import polars as pl

        df = pl.read_csv(GOLDEN_CSV, try_parse_dates=False)
        term_only = df.filter(pl.col("product_type") == "TERM")
        term_csv = tmp_path / "term_only_inforce.csv"
        term_only.write_csv(term_csv)

        xlsx_out = tmp_path / "deal.xlsx"
        result = runner.invoke(
            app,
            [
                "price",
                "--config",
                str(GOLDEN_CONFIGS_DIR / "golden_config_flat.json"),
                "--inforce",
                str(term_csv),
                "--excel-out",
                str(xlsx_out),
            ],
        )
        assert result.exit_code == 0, f"CLI failed:\n{result.stdout}"
        wb = load_workbook(xlsx_out)
        asm = wb["Assumptions"]
        # Map Assumptions rows into a {label: value} dict so we don't
        # depend on the exact row ordering in the writer.
        values = {
            asm.cell(row=r, column=1).value: asm.cell(row=r, column=2).value
            for r in range(3, asm.max_row + 1)
            if asm.cell(row=r, column=1).value is not None
        }
        assert values["Product Type"] == "TERM"
        assert values["Treaty Type"] == "YRT"
        assert values["Cession Percent"] == pytest.approx(0.90)
        assert values["Hurdle Rate"] == pytest.approx(0.10)
        assert values["Discount Rate"] == pytest.approx(0.06)
        assert values["Projection Years"] == 20
