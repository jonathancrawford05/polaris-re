"""
Tests for the deal-pricing Excel export (ADR-045).

The rate-schedule Excel writer is covered separately under
``tests/test_analytics/test_rate_schedule.py``. This module tests the
committee-packet deal-pricing workbook produced by
``write_deal_pricing_excel``.
"""

from datetime import date
from pathlib import Path

import numpy as np
import pytest
from openpyxl import load_workbook

from polaris_re.analytics.alm import DualDurationGap, DurationGapResult
from polaris_re.analytics.ifrs17 import (
    IFRS17CohortManager,
    IFRS17ContractInput,
)
from polaris_re.analytics.premium_sufficiency import PremiumSufficiencyTester
from polaris_re.analytics.profit_test import ProfitResultWithCapital, ProfitTestResult
from polaris_re.core.cashflow import CashFlowResult
from polaris_re.utils.excel_output import (
    AssumptionsMetaExport,
    DealMetaExport,
    DealPricingExport,
    IFRS17MovementExport,
    RatedBlockExport,
    ScenarioMetric,
    write_deal_pricing_excel,
    write_yrt_rate_table_excel,
)

# ---------------------------------------------------------------------------
# Synthetic fixtures
# ---------------------------------------------------------------------------


PROJECTION_YEARS: int = 20
PROJECTION_MONTHS: int = PROJECTION_YEARS * 12


def _make_cashflows(basis: str = "NET", *, scale: float = 1.0) -> CashFlowResult:
    """Build a deterministic CashFlowResult with 240 months of cash flows.

    ``scale`` multiplies every monthly flow so distinct gross / ceded / net
    bases can be built with identifiable magnitudes (used to prove each basis
    lands on its own sheet without cross-wiring).
    """
    t = PROJECTION_MONTHS
    # Monthly arrays — small numbers that aggregate to easy-to-check annual sums.
    premiums = np.full(t, 1_000.0 * scale, dtype=np.float64)  # $12,000 / yr @ scale 1
    claims = np.full(t, 200.0 * scale, dtype=np.float64)  # $2,400 / yr @ scale 1
    surrenders = np.zeros(t, dtype=np.float64)
    expenses = np.full(t, 50.0 * scale, dtype=np.float64)  # $600 / yr @ scale 1
    reserve_balance = np.linspace(0.0, 10_000.0 * scale, t, dtype=np.float64)
    reserve_increase = np.diff(reserve_balance, prepend=0.0)
    ncf = premiums - claims - surrenders - expenses - reserve_increase

    return CashFlowResult(
        run_id="test-run",
        valuation_date=date(2026, 1, 1),
        basis=basis,  # type: ignore[arg-type]
        assumption_set_version="test-v1",
        product_type="TERM",
        projection_months=t,
        gross_premiums=premiums,
        death_claims=claims,
        lapse_surrenders=surrenders,
        expenses=expenses,
        reserve_balance=reserve_balance,
        reserve_increase=reserve_increase,
        net_cash_flow=ncf,
    )


def _make_profit_result(
    *,
    irr: float | None = 0.125,
    profit_margin: float | None = 0.08,
    breakeven_year: int | None = 5,
) -> ProfitTestResult:
    profit_by_year = np.linspace(-1000.0, 5000.0, PROJECTION_YEARS, dtype=np.float64)
    return ProfitTestResult(
        hurdle_rate=0.10,
        pv_profits=25_000.0,
        pv_premiums=250_000.0,
        profit_margin=profit_margin,
        irr=irr,
        breakeven_year=breakeven_year,
        total_undiscounted_profit=float(profit_by_year.sum()),
        profit_by_year=profit_by_year,
    )


def _make_deal_meta() -> DealMetaExport:
    return DealMetaExport(
        product_type="TERM",
        n_policies=500,
        face_amount=50_000_000.0,
        treaty_type="YRT",
        cession_pct=0.90,
        hurdle_rate=0.10,
        discount_rate=0.06,
        projection_years=PROJECTION_YEARS,
        valuation_date=date(2026, 1, 1),
    )


def _make_assumptions_meta() -> AssumptionsMetaExport:
    return AssumptionsMetaExport(
        mortality_source="SOA_VBT_2015",
        mortality_multiplier=1.0,
        lapse_description="Select-ultimate: yr1=10%, yr2=8%, ultimate=3%",
        assumption_set_version="test-v1",
    )


@pytest.fixture
def minimal_export() -> DealPricingExport:
    """Mandatory fields only — cedant + net cash flows, no reinsurer, no scenarios."""
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_profit_result(),
        reinsurer_result=None,
        net_cashflows=_make_cashflows("NET"),
    )


@pytest.fixture
def full_export() -> DealPricingExport:
    """Full export: cedant + reinsurer + scenarios."""
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_profit_result(irr=0.125, profit_margin=0.08, breakeven_year=5),
        reinsurer_result=_make_profit_result(irr=0.095, profit_margin=0.04, breakeven_year=7),
        net_cashflows=_make_cashflows("NET"),
        scenario_results=[
            ScenarioMetric(name="Base", pv_profits=25_000.0, irr=0.125, profit_margin=0.08),
            ScenarioMetric(
                name="Mortality +25%", pv_profits=-12_000.0, irr=0.045, profit_margin=-0.04
            ),
            ScenarioMetric(name="Lapse +50%", pv_profits=18_000.0, irr=0.108, profit_margin=0.06),
        ],
    )


@pytest.fixture
def three_basis_export() -> DealPricingExport:
    """Cedant + reinsurer + scenarios, with distinct gross / ceded / net flows.

    Gross is scale 1.0 (premiums $1,000/mo), ceded scale 0.9 (90% cession ->
    $900/mo), net scale 0.1 ($100/mo). Distinct magnitudes let each sheet be
    checked for its own basis without cross-wiring.
    """
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_profit_result(),
        reinsurer_result=_make_profit_result(irr=0.095, profit_margin=0.04, breakeven_year=7),
        net_cashflows=_make_cashflows("NET", scale=0.1),
        gross_cashflows=_make_cashflows("GROSS", scale=1.0),
        ceded_cashflows=_make_cashflows("CEDED", scale=0.9),
        scenario_results=[
            ScenarioMetric(name="Base", pv_profits=25_000.0, irr=0.125, profit_margin=0.08),
        ],
    )


# ---------------------------------------------------------------------------
# File creation & structure
# ---------------------------------------------------------------------------


class TestDealPricingExcelStructure:
    def test_file_created_and_nonempty(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_expected_sheets_minimal(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        wb = load_workbook(out)
        # Sensitivity omitted when scenario_results is None.
        assert wb.sheetnames == ["Summary", "Cash Flows", "Assumptions"]

    def test_expected_sheets_full(self, full_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        wb = load_workbook(out)
        assert wb.sheetnames == ["Summary", "Cash Flows", "Assumptions", "Sensitivity"]

    def test_workbook_roundtrips(self, full_export: DealPricingExport, tmp_path: Path) -> None:
        """Written workbook can be re-opened with openpyxl without errors."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        wb = load_workbook(out)
        assert wb["Summary"].max_row >= 2
        assert wb["Cash Flows"].max_row >= PROJECTION_YEARS + 1  # header + rows


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------


def _find_row_with_label(ws, label: str) -> int:
    """Return the 1-indexed row containing the given label in column A (None if not found)."""
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1).value
        if cell == label:
            return row_idx
    raise AssertionError(f"Label {label!r} not found in column A of sheet {ws.title!r}")


class TestSummarySheet:
    def test_cedant_irr_matches(self, full_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "IRR")
        # Expect the cedant IRR to live in column B (first metric column).
        cedant_irr_cell = ws.cell(row=row, column=2).value
        assert cedant_irr_cell == pytest.approx(full_export.cedant_result.irr)

    def test_cedant_pv_profits_matches(
        self, full_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "PV Profits")
        pv_cell = ws.cell(row=row, column=2).value
        assert pv_cell == pytest.approx(full_export.cedant_result.pv_profits)

    def test_reinsurer_column_present_when_given(
        self, full_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "IRR")
        reinsurer_cell = ws.cell(row=row, column=3).value
        assert reinsurer_cell == pytest.approx(full_export.reinsurer_result.irr)  # type: ignore[union-attr]

    def test_reinsurer_column_absent_when_none(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Summary"]
        # Row 1 is the header; column B is cedant; column C must not be a reinsurer header.
        header_c = ws.cell(row=1, column=3).value
        assert header_c is None or "Reinsurer" not in str(header_c)

    def test_irr_none_renders_as_na(self, tmp_path: Path) -> None:
        """IRR suppressed by the guardrail (None) must render as 'N/A'."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(irr=None),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "IRR")
        cell = ws.cell(row=row, column=2).value
        assert cell == "N/A"

    def test_profit_margin_none_renders_as_na(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(profit_margin=None),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Profit Margin")
        cell = ws.cell(row=row, column=2).value
        assert cell == "N/A"


# ---------------------------------------------------------------------------
# Cash Flows sheet
# ---------------------------------------------------------------------------


class TestCashFlowsSheet:
    def test_row_count_equals_projection_years(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Cash Flows"]
        # One header row + projection_years data rows.
        assert ws.max_row == PROJECTION_YEARS + 1

    def test_annual_premium_sum_matches_monthly(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Year-1 premium cell = sum of first 12 monthly premiums."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Cash Flows"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        prem_col = headers.index("Gross Premiums") + 1
        year1_prem = ws.cell(row=2, column=prem_col).value
        expected = float(minimal_export.net_cashflows.gross_premiums[:12].sum())
        assert year1_prem == pytest.approx(expected)

    def test_all_expected_columns_present(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Cash Flows"]
        headers = {ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}
        required = {
            "Year",
            "Gross Premiums",
            "Death Claims",
            "Lapse Surrenders",
            "Expenses",
            "Reserve Increase",
            "Net Cash Flow",
        }
        assert required.issubset(headers)

    def test_net_cash_flow_total_matches_profit_sum(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Sum of the annual Net Cash Flow column equals sum of monthly NCF."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Cash Flows"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        ncf_col = headers.index("Net Cash Flow") + 1
        total = sum(
            ws.cell(row=r, column=ncf_col).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=ncf_col).value is not None
        )
        expected = float(minimal_export.net_cashflows.net_cash_flow.sum())
        assert total == pytest.approx(expected, rel=1e-9)


# ---------------------------------------------------------------------------
# Gross / Ceded cash flow sheets
# ---------------------------------------------------------------------------


class TestGrossCededCashFlowSheets:
    """Gross / Ceded cash-flow sheets are written only when their DTO fields
    are populated (purely additive — a None field suppresses the sheet)."""

    def test_sheets_absent_when_fields_none(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        names = load_workbook(out).sheetnames
        assert "Gross Cash Flows" not in names
        assert "Ceded Cash Flows" not in names

    def test_sheets_present_and_ordered_when_populated(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        names = load_workbook(out).sheetnames
        # Gross / Ceded / Net cash-flow block in committee reading order,
        # with the NET sheet keeping its canonical "Cash Flows" title. The
        # combined "Cash Flow Comparison" sheet (ADR-081) follows the NET
        # sheet — written only when both gross and ceded bases are present.
        assert names == [
            "Summary",
            "Gross Cash Flows",
            "Ceded Cash Flows",
            "Cash Flows",
            "Cash Flow Comparison",
            "Line Item Comparison",
            "Assumptions",
            "Sensitivity",
        ]

    def test_gross_sheet_only_when_ceded_none(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET", scale=0.1),
            gross_cashflows=_make_cashflows("GROSS", scale=1.0),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        names = load_workbook(out).sheetnames
        assert names == ["Summary", "Gross Cash Flows", "Cash Flows", "Assumptions"]

    @pytest.mark.parametrize(
        ("sheet", "scale"),
        [
            ("Gross Cash Flows", 1.0),
            ("Ceded Cash Flows", 0.9),
            ("Cash Flows", 0.1),
        ],
    )
    def test_each_sheet_carries_its_own_basis(
        self,
        three_basis_export: DealPricingExport,
        tmp_path: Path,
        sheet: str,
        scale: float,
    ) -> None:
        """Year-1 Gross Premiums cell on each sheet equals that basis' own
        annual premium sum — proves no cross-wiring between bases."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        ws = load_workbook(out)[sheet]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        prem_col = headers.index("Gross Premiums") + 1
        year1_prem = ws.cell(row=2, column=prem_col).value
        expected = 12 * 1_000.0 * scale  # 12 months at $1,000 * scale
        assert year1_prem == pytest.approx(expected)

    def test_gross_and_net_premiums_differ(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Sanity: gross premiums exceed net premiums on the rendered sheets."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        wb = load_workbook(out)

        def _year1_premium(sheet: str) -> float:
            ws = wb[sheet]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            prem_col = headers.index("Gross Premiums") + 1
            return ws.cell(row=2, column=prem_col).value

        assert _year1_premium("Gross Cash Flows") > _year1_premium("Cash Flows")


# ---------------------------------------------------------------------------
# Combined Gross / Ceded / Net cash-flow comparison sheet (ADR-081)
# ---------------------------------------------------------------------------


class TestCashFlowComparisonSheet:
    """The combined "Cash Flow Comparison" sheet places the per-year Net Cash
    Flow of all three bases side by side with a ``Gross - Ceded`` check column.
    It is written only when BOTH gross and ceded bases are populated (the
    comparison is meaningless with a missing basis), so net-only and
    gross-only exports stay byte-identical."""

    def test_absent_when_net_only(self, minimal_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        assert "Cash Flow Comparison" not in load_workbook(out).sheetnames

    def test_absent_when_ceded_missing(self, tmp_path: Path) -> None:
        """Gross present but ceded None — no comparison sheet (incomplete bases)."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET", scale=0.1),
            gross_cashflows=_make_cashflows("GROSS", scale=1.0),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        assert "Cash Flow Comparison" not in load_workbook(out).sheetnames

    def test_present_when_all_three_bases(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        assert "Cash Flow Comparison" in load_workbook(out).sheetnames

    def test_columns(self, three_basis_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        ws = load_workbook(out)["Cash Flow Comparison"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        assert headers == ["Year", "Gross", "Ceded", "Net", "Gross - Ceded"]

    def test_row_count_equals_projection_years(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        ws = load_workbook(out)["Cash Flow Comparison"]
        assert ws.max_row == PROJECTION_YEARS + 1  # header + one row per year

    def test_columns_match_basis_sheets(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Each basis column equals that basis sheet's own annual Net Cash Flow."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        wb = load_workbook(out)
        comp = wb["Cash Flow Comparison"]

        def _ncf_column(sheet: str) -> list[float]:
            ws = wb[sheet]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            col = headers.index("Net Cash Flow") + 1
            return [ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)]

        gross_basis = _ncf_column("Gross Cash Flows")
        ceded_basis = _ncf_column("Ceded Cash Flows")
        net_basis = _ncf_column("Cash Flows")
        for i, r in enumerate(range(2, comp.max_row + 1)):
            assert comp.cell(row=r, column=2).value == pytest.approx(gross_basis[i])
            assert comp.cell(row=r, column=3).value == pytest.approx(ceded_basis[i])
            assert comp.cell(row=r, column=4).value == pytest.approx(net_basis[i])

    def test_gross_minus_ceded_identity(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Closed-form: the ``Gross - Ceded`` column equals both the arithmetic
        difference of the Gross/Ceded columns and the Net column per year
        (Net = Gross - Ceded — the core treaty decomposition)."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        ws = load_workbook(out)["Cash Flow Comparison"]
        for r in range(2, ws.max_row + 1):
            gross = ws.cell(row=r, column=2).value
            ceded = ws.cell(row=r, column=3).value
            net = ws.cell(row=r, column=4).value
            check = ws.cell(row=r, column=5).value
            assert check == pytest.approx(gross - ceded)
            assert check == pytest.approx(net)


# ---------------------------------------------------------------------------
# Per-line-item Gross / Ceded / Net comparison sheet (ADR-086)
# ---------------------------------------------------------------------------


# Component cash-flow line items broken out on the Line Item Comparison sheet
# (every basis-sheet column except "Year" and the bottom-line "Net Cash Flow",
# which the Cash Flow Comparison sheet already diffs).
_LINE_ITEMS: tuple[str, ...] = (
    "Gross Premiums",
    "Death Claims",
    "Lapse Surrenders",
    "Expenses",
    "Reserve Increase",
)


class TestLineItemComparisonSheet:
    """The "Line Item Comparison" sheet places each component cash-flow line
    item (premiums, claims, surrenders, expenses, reserve increase) of all
    three bases side by side, so a committee can see where the ceded share
    concentrates rather than only the bottom-line Net Cash Flow. It is written
    only when BOTH gross and ceded bases are populated (the same gate as the
    Cash Flow Comparison sheet), so net-only / gross-only exports stay
    byte-identical."""

    def test_absent_when_net_only(self, minimal_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        assert "Line Item Comparison" not in load_workbook(out).sheetnames

    def test_absent_when_ceded_missing(self, tmp_path: Path) -> None:
        """Gross present but ceded None — no line-item sheet (incomplete bases)."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET", scale=0.1),
            gross_cashflows=_make_cashflows("GROSS", scale=1.0),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        assert "Line Item Comparison" not in load_workbook(out).sheetnames

    def test_present_and_ordered_when_all_three_bases(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        names = load_workbook(out).sheetnames
        assert "Line Item Comparison" in names
        # Follows the bottom-line Cash Flow Comparison sheet immediately.
        assert names.index("Line Item Comparison") == names.index("Cash Flow Comparison") + 1

    def test_columns(self, three_basis_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        ws = load_workbook(out)["Line Item Comparison"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        expected = ["Year"]
        for item in _LINE_ITEMS:
            expected += [f"{item} (Gross)", f"{item} (Ceded)", f"{item} (Net)"]
        assert headers == expected

    def test_row_count_equals_projection_years(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        ws = load_workbook(out)["Line Item Comparison"]
        assert ws.max_row == PROJECTION_YEARS + 1  # header + one row per year

    def test_columns_match_basis_sheets(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Each (line item, basis) triplet column equals that basis sheet's own
        annual value for the same line item — proves no cross-wiring."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        wb = load_workbook(out)
        comp = wb["Line Item Comparison"]
        comp_headers = [comp.cell(row=1, column=c).value for c in range(1, comp.max_column + 1)]

        def _basis_column(sheet: str, item: str) -> list[float]:
            ws = wb[sheet]
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            col = headers.index(item) + 1
            return [ws.cell(row=r, column=col).value for r in range(2, ws.max_row + 1)]

        for item in _LINE_ITEMS:
            for basis_label, sheet in (
                ("Gross", "Gross Cash Flows"),
                ("Ceded", "Ceded Cash Flows"),
                ("Net", "Cash Flows"),
            ):
                comp_col = comp_headers.index(f"{item} ({basis_label})") + 1
                basis_vals = _basis_column(sheet, item)
                for i, r in enumerate(range(2, comp.max_row + 1)):
                    assert comp.cell(row=r, column=comp_col).value == pytest.approx(basis_vals[i])

    def test_net_equals_gross_minus_ceded_per_line_item(
        self, three_basis_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Closed-form: for every line item and every year the Net column equals
        Gross - Ceded (the treaty decomposition holds component-by-component,
        not only on the bottom-line Net Cash Flow)."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(three_basis_export, out)
        ws = load_workbook(out)["Line Item Comparison"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for item in _LINE_ITEMS:
            g_col = headers.index(f"{item} (Gross)") + 1
            c_col = headers.index(f"{item} (Ceded)") + 1
            n_col = headers.index(f"{item} (Net)") + 1
            for r in range(2, ws.max_row + 1):
                gross = ws.cell(row=r, column=g_col).value
                ceded = ws.cell(row=r, column=c_col).value
                net = ws.cell(row=r, column=n_col).value
                assert net == pytest.approx(gross - ceded)


# ---------------------------------------------------------------------------
# Assumptions sheet
# ---------------------------------------------------------------------------


class TestAssumptionsSheet:
    def test_contains_mortality_source(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Mortality Source" in labels

    def test_contains_treaty_and_cession(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Treaty Type" in labels
        assert "Cession Percent" in labels

    def test_hurdle_rate_value(self, minimal_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Assumptions"]
        row = _find_row_with_label(ws, "Hurdle Rate")
        assert ws.cell(row=row, column=2).value == pytest.approx(
            minimal_export.deal_meta.hurdle_rate
        )

    def test_reserve_basis_row_defaults_to_net_premium(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """The Assumptions sheet always labels the reserve basis (reserve-basis epic)."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Assumptions"]
        row = _find_row_with_label(ws, "Reserve Basis")
        assert row is not None
        assert ws.cell(row=row, column=2).value == "NET_PREMIUM"

    def test_reserve_basis_row_reflects_non_default(self, tmp_path: Path) -> None:
        """A non-default basis on the DealMetaExport is rendered verbatim."""
        from dataclasses import replace

        export = DealPricingExport(
            deal_meta=replace(_make_deal_meta(), reserve_basis="CRVM"),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Assumptions"]
        row = _find_row_with_label(ws, "Reserve Basis")
        assert ws.cell(row=row, column=2).value == "CRVM"


# ---------------------------------------------------------------------------
# Rated-block panel on the Assumptions sheet (ADR-068)
# ---------------------------------------------------------------------------


def _make_rated_block(
    *,
    n_policies: int = 500,
    n_rated: int = 60,
    pct_rated_by_count: float = 0.12,
    pct_rated_by_face: float = 0.18,
    face_weighted_mean_multiplier: float = 1.075,
    max_multiplier: float = 3.0,
    max_flat_extra_per_1000: float = 5.0,
) -> RatedBlockExport:
    return RatedBlockExport(
        n_policies=n_policies,
        n_rated=n_rated,
        pct_rated_by_count=pct_rated_by_count,
        pct_rated_by_face=pct_rated_by_face,
        face_weighted_mean_multiplier=face_weighted_mean_multiplier,
        max_multiplier=max_multiplier,
        max_flat_extra_per_1000=max_flat_extra_per_1000,
    )


class TestRatedBlockPanel:
    """Open Question #3 from CONTINUATION_deal_pricing_excel (ADR-068).

    The Assumptions sheet picks up an optional panel of substandard-rating
    composition. Suppressed when ``rated_block is None`` or when no rated
    lives are present, so all-standard blocks remain byte-identical to
    pre-ADR-068 output.
    """

    def test_panel_absent_by_default(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Rated Block" not in labels
        assert "Policies Rated" not in labels

    def test_panel_suppressed_when_n_rated_zero(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            rated_block=_make_rated_block(
                n_rated=0,
                pct_rated_by_count=0.0,
                pct_rated_by_face=0.0,
                face_weighted_mean_multiplier=1.0,
                max_multiplier=1.0,
                max_flat_extra_per_1000=0.0,
            ),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Policies Rated" not in labels

    def test_panel_renders_when_rated_lives_present(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            rated_block=_make_rated_block(),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        # Section header + the six labelled rows from `rating_composition`.
        assert "Rated Block" in labels
        assert "Policies Rated" in labels
        assert "% Rated (by count)" in labels
        assert "% Rated (by face)" in labels
        assert "Face-weighted Avg Multiplier" in labels
        assert "Max Multiplier" in labels
        assert "Max Flat Extra / $1,000" in labels

    def test_panel_n_rated_value(self, tmp_path: Path) -> None:
        rated = _make_rated_block(n_rated=73)
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            rated_block=rated,
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Assumptions"]
        row = _find_row_with_label(ws, "Policies Rated")
        assert ws.cell(row=row, column=2).value == 73

    def test_panel_face_weighted_multiplier_value(self, tmp_path: Path) -> None:
        rated = _make_rated_block(face_weighted_mean_multiplier=1.234)
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            rated_block=rated,
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Assumptions"]
        row = _find_row_with_label(ws, "Face-weighted Avg Multiplier")
        assert ws.cell(row=row, column=2).value == pytest.approx(1.234)

    def test_panel_percentage_formatting(self, tmp_path: Path) -> None:
        """`pct_rated_*` cells render with a percent number format."""
        rated = _make_rated_block(pct_rated_by_count=0.123, pct_rated_by_face=0.21)
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            rated_block=rated,
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Assumptions"]
        row = _find_row_with_label(ws, "% Rated (by count)")
        cell = ws.cell(row=row, column=2)
        assert cell.value == pytest.approx(0.123)
        assert "%" in cell.number_format


# ---------------------------------------------------------------------------
# Sensitivity sheet
# ---------------------------------------------------------------------------


class TestSensitivitySheet:
    def test_row_per_scenario(self, full_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        ws = load_workbook(out)["Sensitivity"]
        assert full_export.scenario_results is not None
        # header + one row per scenario
        assert ws.max_row == len(full_export.scenario_results) + 1

    def test_scenario_names_preserved(self, full_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        ws = load_workbook(out)["Sensitivity"]
        names_in_sheet = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert full_export.scenario_results is not None
        assert names_in_sheet == [s.name for s in full_export.scenario_results]

    def test_scenario_pv_matches(self, full_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(full_export, out)
        ws = load_workbook(out)["Sensitivity"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        pv_col = headers.index("PV Profits") + 1
        assert full_export.scenario_results is not None
        for i, scenario in enumerate(full_export.scenario_results):
            pv_cell = ws.cell(row=i + 2, column=pv_col).value
            assert pv_cell == pytest.approx(scenario.pv_profits)


# ---------------------------------------------------------------------------
# LICAT capital block on the Summary sheet (ADR-049)
# ---------------------------------------------------------------------------


def _make_capital_result(
    *,
    pv_capital: float = 100_000.0,
    pv_capital_strain: float = 25_000.0,
    return_on_capital: float | None = 0.18,
    capital_adjusted_irr: float | None = 0.115,
    peak_capital: float = 150_000.0,
    available_capital: float | None = None,
    capital_ratio: float | None = None,
) -> ProfitResultWithCapital:
    profit_by_year = np.linspace(-1000.0, 5000.0, PROJECTION_YEARS, dtype=np.float64)
    return ProfitResultWithCapital(
        hurdle_rate=0.10,
        pv_profits=25_000.0,
        pv_premiums=250_000.0,
        profit_margin=0.10,
        irr=0.125,
        breakeven_year=5,
        total_undiscounted_profit=float(profit_by_year.sum()),
        profit_by_year=profit_by_year,
        initial_capital=80_000.0,
        peak_capital=peak_capital,
        pv_capital=pv_capital,
        pv_capital_strain=pv_capital_strain,
        return_on_capital=return_on_capital,
        capital_adjusted_irr=capital_adjusted_irr,
        capital_by_period=np.linspace(80_000.0, peak_capital, PROJECTION_MONTHS, dtype=np.float64),
        available_capital=available_capital,
        capital_ratio=capital_ratio,
    )


@pytest.fixture
def capital_export() -> DealPricingExport:
    """Export with a ProfitResultWithCapital cedant + reinsurer."""
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_capital_result(),
        reinsurer_result=_make_capital_result(
            pv_capital=60_000.0,
            return_on_capital=0.22,
            capital_adjusted_irr=0.142,
            peak_capital=90_000.0,
        ),
        net_cashflows=_make_cashflows("NET"),
    )


class TestSummarySheetCapitalBlock:
    """ADR-049: capital rows are appended only when results carry capital."""

    def test_capital_rows_absent_when_off(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Return on Capital" not in labels
        assert "Peak Capital" not in labels
        assert "PV Capital (stock)" not in labels
        assert "PV Capital Strain" not in labels
        assert "Capital-Adjusted IRR" not in labels

    def test_capital_rows_present_when_capital_result(
        self, capital_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(capital_export, out)
        ws = load_workbook(out)["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        for required in (
            "Peak Capital",
            "PV Capital (stock)",
            "PV Capital Strain",
            "Return on Capital",
            "Capital-Adjusted IRR",
        ):
            assert required in labels, f"missing {required!r} on Summary sheet"

    def test_cedant_return_on_capital_value_matches(
        self, capital_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(capital_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Return on Capital")
        cell = ws.cell(row=row, column=2).value
        assert cell == pytest.approx(0.18)

    def test_reinsurer_pv_capital_value_matches(
        self, capital_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(capital_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "PV Capital (stock)")
        cedant_cell = ws.cell(row=row, column=2).value
        reinsurer_cell = ws.cell(row=row, column=3).value
        assert cedant_cell == pytest.approx(100_000.0)
        assert reinsurer_cell == pytest.approx(60_000.0)

    def test_advisory_strain_metric_present(
        self, capital_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """PR-#34 reviewer requirement: surface PV Capital Strain on Excel."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(capital_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "PV Capital Strain")
        cedant_cell = ws.cell(row=row, column=2).value
        assert cedant_cell == pytest.approx(25_000.0)

    def test_return_on_capital_none_renders_as_na(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_capital_result(return_on_capital=None),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Return on Capital")
        assert ws.cell(row=row, column=2).value == "N/A"

    def test_reinsurer_capital_na_when_only_cedant_has_capital(self, tmp_path: Path) -> None:
        """Mixed run: cedant has capital, reinsurer is plain — reinsurer cells = N/A."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_capital_result(),
            reinsurer_result=_make_profit_result(),  # plain ProfitTestResult
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Return on Capital")
        assert ws.cell(row=row, column=3).value == "N/A"


class TestCapitalJurisdictionHeader:
    """ADR-102 (Slice 4b): the capital block names the regulatory standard."""

    def _labels(self, export: DealPricingExport, tmp_path: Path) -> list[object]:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        return [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]

    @pytest.mark.parametrize(
        ("model_id", "expected"),
        [
            ("licat", "Regulatory Capital — LICAT (Canada)"),
            ("rbc", "Regulatory Capital — US RBC"),
            ("solvency2", "Regulatory Capital — EU Solvency II"),
        ],
    )
    def test_header_names_jurisdiction(
        self,
        capital_export: DealPricingExport,
        tmp_path: Path,
        model_id: str,
        expected: str,
    ) -> None:
        import dataclasses

        export = dataclasses.replace(capital_export, capital_model_id=model_id)
        labels = self._labels(export, tmp_path)
        assert expected in labels
        # The header sits directly above the first capital metric.
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        assert _find_row_with_label(ws, expected) + 1 == _find_row_with_label(ws, "Peak Capital")

    def test_none_defaults_to_licat_label(
        self, capital_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """An un-tagged export (capital_model_id=None) is LICAT by construction."""
        labels = self._labels(capital_export, tmp_path)  # fixture leaves id None
        assert "Regulatory Capital — LICAT (Canada)" in labels

    def test_header_absent_when_no_capital(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        labels = self._labels(minimal_export, tmp_path)
        assert not any(
            isinstance(value, str) and value.startswith("Regulatory Capital") for value in labels
        )


class TestSummarySheetSolvencyRatio:
    """ADR-104 / Slice 4c-2b: the capital block gains Available Capital +

    Solvency Ratio rows only when an available-capital numerator was supplied.
    """

    def _labels(self, export: DealPricingExport, tmp_path: Path) -> list[object]:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        return [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]

    def test_ratio_rows_absent_when_capital_run_has_no_numerator(
        self, capital_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """A capital run WITHOUT --available-capital is byte-identical (no rows)."""
        labels = self._labels(capital_export, tmp_path)  # fixture leaves ratio None
        assert "Solvency Ratio" not in labels
        assert "Available Capital" not in labels

    def test_ratio_rows_absent_when_no_capital(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        labels = self._labels(minimal_export, tmp_path)
        assert "Solvency Ratio" not in labels
        assert "Available Capital" not in labels

    def test_ratio_rows_present_when_numerator_supplied(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_capital_result(available_capital=2_000_000.0, capital_ratio=2.5),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            capital_model_id="rbc",
        )
        labels = self._labels(export, tmp_path)
        assert "Available Capital" in labels
        assert "Solvency Ratio" in labels

    def test_ratio_rows_sit_below_capital_block(self, tmp_path: Path) -> None:
        """The two ratio rows follow the last standing capital metric."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_capital_result(available_capital=2_000_000.0, capital_ratio=2.5),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        last_metric = _find_row_with_label(ws, "Capital-Adjusted IRR")
        avail_row = _find_row_with_label(ws, "Available Capital")
        ratio_row = _find_row_with_label(ws, "Solvency Ratio")
        assert avail_row == last_metric + 1
        assert ratio_row == avail_row + 1

    def test_cedant_ratio_and_numerator_values_match(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_capital_result(available_capital=2_000_000.0, capital_ratio=2.5),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        avail = ws.cell(row=_find_row_with_label(ws, "Available Capital"), column=2).value
        ratio = ws.cell(row=_find_row_with_label(ws, "Solvency Ratio"), column=2).value
        assert avail == pytest.approx(2_000_000.0)
        assert ratio == pytest.approx(2.5)

    def test_reinsurer_ratio_na_when_only_cedant_has_numerator(self, tmp_path: Path) -> None:
        """Mixed run: cedant carries the ratio, reinsurer is plain -> reinsurer cells N/A."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_capital_result(available_capital=2_000_000.0, capital_ratio=2.5),
            reinsurer_result=_make_profit_result(),  # plain ProfitTestResult
            net_cashflows=_make_cashflows("NET"),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        ratio_row = _find_row_with_label(ws, "Solvency Ratio")
        assert ws.cell(row=ratio_row, column=3).value == "N/A"


# ---------------------------------------------------------------------------
# YRT Rate Table sheet (ADR-052)
# ---------------------------------------------------------------------------


def _make_yrt_rate_table():  # type: ignore[no-untyped-def]
    """Build a small synthetic ``YRTRateTable`` for sheet-rendering tests."""
    from polaris_re.core.policy import Sex, SmokerStatus
    from polaris_re.reinsurance.yrt_rate_table import YRTRateTable, YRTRateTableArray

    rates_male_ns = np.array(
        [
            [0.50, 0.55, 0.60, 1.00],
            [0.55, 0.60, 0.65, 1.10],
            [0.60, 0.65, 0.70, 1.20],
        ],
        dtype=np.float64,
    )
    rates_male_smoker = rates_male_ns * 1.6
    arr_ns = YRTRateTableArray(rates=rates_male_ns, min_age=30, max_age=32, select_period=3)
    arr_smoker = YRTRateTableArray(rates=rates_male_smoker, min_age=30, max_age=32, select_period=3)
    return YRTRateTable.from_arrays(
        table_name="synthetic-test",
        arrays={
            (Sex.MALE, SmokerStatus.NON_SMOKER): arr_ns,
            (Sex.MALE, SmokerStatus.SMOKER): arr_smoker,
        },
    )


class TestYRTRateTableSheet:
    """`write_deal_pricing_excel` adds a ``YRT Rate Table`` sheet when set."""

    def test_omitted_when_yrt_rate_table_is_none(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        wb = load_workbook(out)
        assert "YRT Rate Table" not in wb.sheetnames

    def test_added_when_yrt_rate_table_set(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            yrt_rate_table=_make_yrt_rate_table(),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        wb = load_workbook(out)
        assert "YRT Rate Table" in wb.sheetnames

    def test_sheet_contains_table_name_and_cohorts(self, tmp_path: Path) -> None:
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            yrt_rate_table=_make_yrt_rate_table(),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["YRT Rate Table"]
        # Title (row 1) names the table.
        assert "synthetic-test" in str(ws.cell(row=1, column=1).value)
        # Find both cohort labels in column A.
        labels = [str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)]
        assert any("M_NS" in lbl for lbl in labels)
        assert any("M_S" in lbl for lbl in labels)

    def test_sheet_renders_known_rate_value(self, tmp_path: Path) -> None:
        """Smoke check: at least one cell holds the expected rate value."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            yrt_rate_table=_make_yrt_rate_table(),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["YRT Rate Table"]
        # Walk every cell — at least one should equal one of our rates.
        target = 0.50  # M_NS, age 30, dur_1
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and abs(cell.value - target) < 1e-9:
                    found = True
                    break
            if found:
                break
        assert found


# ---------------------------------------------------------------------------
# write_yrt_rate_table_excel — standalone workbook (ADR-053)
# ---------------------------------------------------------------------------


class TestWriteYrtRateTableExcel:
    """``write_yrt_rate_table_excel`` produces a standalone workbook."""

    def test_workbook_created(self, tmp_path: Path) -> None:
        """Workbook file is written and non-empty."""
        table = _make_yrt_rate_table()
        out = tmp_path / "schedule.xlsx"
        write_yrt_rate_table_excel(table, out)
        assert out.exists()
        assert out.stat().st_size > 0

    def test_has_summary_and_rate_table_sheets(self, tmp_path: Path) -> None:
        """Workbook carries both ``Summary`` and ``YRT Rate Table`` sheets."""
        table = _make_yrt_rate_table()
        out = tmp_path / "schedule.xlsx"
        write_yrt_rate_table_excel(table, out)
        wb = load_workbook(out)
        assert "Summary" in wb.sheetnames
        assert "YRT Rate Table" in wb.sheetnames

    def test_summary_carries_table_metadata(self, tmp_path: Path) -> None:
        """``Summary`` sheet shows the table name and cohort count."""
        table = _make_yrt_rate_table()
        out = tmp_path / "schedule.xlsx"
        write_yrt_rate_table_excel(table, out)
        ws = load_workbook(out)["Summary"]
        cells = [str(ws.cell(row=r, column=1).value) for r in range(1, 10)]
        assert any("synthetic-test" in c for c in cells)
        # Two cohorts in the fixture.
        assert any("Cohorts: 2" in c for c in cells)
        # Total rate cells: 2 cohorts x 3 ages x 4 cols = 24.
        assert any("Total rate cells: 24" in c for c in cells)

    def test_rate_table_sheet_renders_known_value(self, tmp_path: Path) -> None:
        """At least one cell carries the expected M_NS / age 30 / dur_1 rate."""
        table = _make_yrt_rate_table()
        out = tmp_path / "schedule.xlsx"
        write_yrt_rate_table_excel(table, out)
        ws = load_workbook(out)["YRT Rate Table"]
        target = 0.50  # M_NS, age 30, dur_1
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and abs(cell.value - target) < 1e-9:
                    found = True
                    break
            if found:
                break
        assert found

    def test_workbook_omits_empty_default_sheet(self, tmp_path: Path) -> None:
        """openpyxl's default ``Sheet`` is removed before adding the named sheets."""
        table = _make_yrt_rate_table()
        out = tmp_path / "schedule.xlsx"
        write_yrt_rate_table_excel(table, out)
        wb = load_workbook(out)
        assert "Sheet" not in wb.sheetnames
        # Only the two intentional sheets.
        assert set(wb.sheetnames) == {"Summary", "YRT Rate Table"}

    def test_wide_select_period_column_widths(self, tmp_path: Path) -> None:
        """`select_period >= 25` does not crash or corrupt column widths.

        Regression for PR #39 P1: ``chr(ord("B") + col_offset)`` produced
        invalid Excel column keys (`'['`, `'\\\\'`, ...) for select periods
        past 24, even though `YRTRateTable.select_period_years` allows up
        to 50. The fix uses ``openpyxl.utils.get_column_letter`` which
        wraps to two-letter columns (`AA`, `AB`, ...) past Z.
        """
        from polaris_re.core.policy import Sex, SmokerStatus
        from polaris_re.reinsurance.yrt_rate_table import YRTRateTable, YRTRateTableArray

        # select_period=30 → 31 duration columns (dur_1 .. ultimate),
        # data columns B .. AF (need two-letter columns past Z).
        select_period = 30
        rates = np.full((2, select_period + 1), 0.5, dtype=np.float64)
        arr = YRTRateTableArray(rates=rates, min_age=30, max_age=31, select_period=select_period)
        table = YRTRateTable.from_arrays(
            table_name="wide-select",
            arrays={(Sex.MALE, SmokerStatus.NON_SMOKER): arr},
        )
        out = tmp_path / "wide.xlsx"
        write_yrt_rate_table_excel(table, out)
        wb = load_workbook(out)
        ws = wb["YRT Rate Table"]
        # Column B holds dur_1; column AF (= index 32) holds the ultimate.
        # Both must have the explicit width set (14.0) by the loop.
        assert ws.column_dimensions["B"].width == 14
        assert ws.column_dimensions["AF"].width == 14
        # And no corrupt non-letter keys (e.g. '[') made it in.
        for key in ws.column_dimensions:
            assert key.isalpha(), f"corrupt column key {key!r}"


# ---------------------------------------------------------------------------
# Filled-cell disclosure (ADR-054) — visual styling + Summary count
# ---------------------------------------------------------------------------


def _make_partially_solved_table():
    """Two-row cohort table with the second row marked as filled."""
    from polaris_re.core.policy import Sex, SmokerStatus
    from polaris_re.reinsurance.yrt_rate_table import YRTRateTable, YRTRateTableArray

    rates = np.array([[1.0], [2.0]], dtype=np.float64)
    mask = np.array([[True], [False]], dtype=np.bool_)
    arr = YRTRateTableArray(
        rates=rates,
        min_age=40,
        max_age=41,
        select_period=0,
        solved_mask=mask,
    )
    return YRTRateTable.from_arrays(
        table_name="partial",
        arrays={(Sex.MALE, SmokerStatus.UNKNOWN): arr},
    )


class TestSolvedMaskDisclosureExcel:
    """ADR-054 — Excel surfaces disclose forward/back-filled cells."""

    def test_filled_cell_is_italic(self, tmp_path: Path) -> None:
        """Cells with mask False render in italic font."""
        table = _make_partially_solved_table()
        out = tmp_path / "partial.xlsx"
        write_yrt_rate_table_excel(table, out)
        ws = load_workbook(out)["YRT Rate Table"]
        # Find the two data cells. Layout per ``_write_yrt_rate_table_sheet``:
        # title row 1, age-range row 2, NOTE row 3 (because mask present),
        # blank row 4, cohort label row 5, header row 6, data rows 7 & 8.
        solved_cell = ws["B7"]
        filled_cell = ws["B8"]
        assert solved_cell.value == 1.0
        assert filled_cell.value == 2.0
        assert solved_cell.font.italic is False
        assert filled_cell.font.italic is True

    def test_filled_cell_has_grey_fill(self, tmp_path: Path) -> None:
        """Cells with mask False render with a light-grey ``PatternFill``."""
        table = _make_partially_solved_table()
        out = tmp_path / "partial.xlsx"
        write_yrt_rate_table_excel(table, out)
        ws = load_workbook(out)["YRT Rate Table"]
        filled_cell = ws["B8"]
        # ``EEEEEE`` (defined in `_write_yrt_rate_table_sheet`) is the disclosure colour.
        # openpyxl serialises it as ``00EEEEEE`` (alpha-prefixed), so check the suffix.
        fg = filled_cell.fill.fgColor.rgb if filled_cell.fill.fgColor is not None else ""
        assert "EEEEEE" in (fg or "").upper()

    def test_disclosure_note_row_present(self, tmp_path: Path) -> None:
        """Row 3 carries the human-readable note when any cohort has filled cells."""
        table = _make_partially_solved_table()
        out = tmp_path / "partial.xlsx"
        write_yrt_rate_table_excel(table, out)
        ws = load_workbook(out)["YRT Rate Table"]
        note = str(ws.cell(row=3, column=1).value or "")
        assert "forward/back-filled" in note
        assert "ADR-054" in note

    def test_summary_carries_solved_filled_counts(self, tmp_path: Path) -> None:
        """The Summary sheet records solved- and filled-cell counts when masked."""
        table = _make_partially_solved_table()
        out = tmp_path / "partial.xlsx"
        write_yrt_rate_table_excel(table, out)
        ws = load_workbook(out)["Summary"]
        cells = [str(ws.cell(row=r, column=1).value or "") for r in range(1, 12)]
        assert any("Solved cells: 1" in c for c in cells)
        assert any("Filled cells: 1" in c for c in cells)

    def test_no_mask_renders_unchanged(self, tmp_path: Path) -> None:
        """CSV-loaded tables (mask None) keep the pre-ADR-054 layout.

        - Title at row 1
        - Age-range row at row 2
        - NO note row at row 3
        - Cohort label at row 4
        - No italic / no light-grey fill on data cells
        - Summary lacks the Solved / Filled count rows
        """
        from polaris_re.core.policy import Sex, SmokerStatus
        from polaris_re.reinsurance.yrt_rate_table import YRTRateTable, YRTRateTableArray

        rates = np.array([[1.0], [2.0]], dtype=np.float64)
        arr = YRTRateTableArray(rates=rates, min_age=40, max_age=41, select_period=0)
        table = YRTRateTable.from_arrays(
            table_name="loaded",
            arrays={(Sex.MALE, SmokerStatus.UNKNOWN): arr},
        )
        out = tmp_path / "loaded.xlsx"
        write_yrt_rate_table_excel(table, out)

        wb = load_workbook(out)
        ws = wb["YRT Rate Table"]
        # Cohort label sits at the original row 4 (no NOTE row inserted).
        assert "Cohort:" in str(ws.cell(row=4, column=1).value)
        # Data rows at 6 & 7 (header at 5).
        c1 = ws["B6"]
        c2 = ws["B7"]
        assert c1.font.italic is False
        assert c2.font.italic is False
        # No grey fill applied.
        for cell in (c1, c2):
            fg = cell.fill.fgColor.rgb if cell.fill.fgColor is not None else None
            assert (fg is None) or ("EEEEEE" not in (fg or "").upper())

        ws_summary = wb["Summary"]
        cells = [str(ws_summary.cell(row=r, column=1).value or "") for r in range(1, 12)]
        # The Solved / Filled disclosure is gated on mask presence — must NOT appear.
        assert not any("Solved cells:" in c for c in cells)
        assert not any("Filled cells:" in c for c in cells)


# ---------------------------------------------------------------------------
# Premium-sufficiency panel (ADR-083)
# ---------------------------------------------------------------------------


def _make_sufficiency(scale: float = 1.0, *, target_margin: float = 0.0):
    """Run the analyzer on a deterministic GROSS cash flow for panel tests."""
    cf = _make_cashflows("GROSS", scale=scale)
    return PremiumSufficiencyTester(cf, discount_rate=0.06, target_margin=target_margin).run()


@pytest.fixture
def sufficiency_export() -> DealPricingExport:
    """Full export carrying cedant + reinsurer premium-sufficiency results."""
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_profit_result(),
        reinsurer_result=_make_profit_result(irr=0.095, profit_margin=0.04, breakeven_year=7),
        net_cashflows=_make_cashflows("NET"),
        premium_sufficiency_cedant=_make_sufficiency(scale=1.0, target_margin=0.05),
        premium_sufficiency_reinsurer=_make_sufficiency(scale=0.9, target_margin=0.05),
    )


class TestPremiumSufficiencyPanel:
    _ROWS = (
        "Sufficiency Discount Rate",
        "Sufficiency Target Margin",
        "PV Claims",
        "PV Surrenders",
        "PV Benefits",
        "PV Expenses",
        "Sufficiency Margin",
        "Loss Ratio",
        "Expense Ratio",
        "Combined Ratio",
        "Premium Sufficient",
    )

    def test_panel_absent_when_not_populated(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """Backward compat: no sufficiency data -> no panel rows."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        for row in self._ROWS:
            assert row not in labels

    def test_panel_rows_present_when_populated(
        self, sufficiency_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(sufficiency_export, out)
        ws = load_workbook(out)["Summary"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        for row in self._ROWS:
            assert row in labels, f"missing {row!r} on Summary sheet"

    def test_cedant_combined_ratio_cell_matches(
        self, sufficiency_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(sufficiency_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Combined Ratio")
        cell = ws.cell(row=row, column=2).value
        assert cell == pytest.approx(sufficiency_export.premium_sufficiency_cedant.combined_ratio)

    def test_sufficiency_margin_cell_matches(
        self, sufficiency_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(sufficiency_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Sufficiency Margin")
        cell = ws.cell(row=row, column=2).value
        assert cell == pytest.approx(
            sufficiency_export.premium_sufficiency_cedant.sufficiency_margin
        )

    def test_pv_claims_and_surrenders_cells_match(
        self, sufficiency_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """ADR-084: the PV Claims / PV Surrenders breakdown rows carry the
        analyzer's per-line-item components."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(sufficiency_export, out)
        ws = load_workbook(out)["Summary"]
        result = sufficiency_export.premium_sufficiency_cedant
        claims_row = _find_row_with_label(ws, "PV Claims")
        surr_row = _find_row_with_label(ws, "PV Surrenders")
        assert ws.cell(row=claims_row, column=2).value == pytest.approx(result.pv_claims)
        assert ws.cell(row=surr_row, column=2).value == pytest.approx(result.pv_surrenders)

    def test_claims_plus_surrenders_equals_benefits(
        self, sufficiency_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """ADR-084 closed-form: the two breakdown rows sum to the PV Benefits
        row on the sheet (PV Benefits = PV Claims + PV Surrenders)."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(sufficiency_export, out)
        ws = load_workbook(out)["Summary"]
        claims = ws.cell(row=_find_row_with_label(ws, "PV Claims"), column=2).value
        surrenders = ws.cell(row=_find_row_with_label(ws, "PV Surrenders"), column=2).value
        benefits = ws.cell(row=_find_row_with_label(ws, "PV Benefits"), column=2).value
        assert claims + surrenders == pytest.approx(benefits)

    def test_verdict_cell_is_yes_or_no(
        self, sufficiency_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(sufficiency_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Premium Sufficient")
        expected = "Yes" if sufficiency_export.premium_sufficiency_cedant.is_sufficient else "No"
        assert ws.cell(row=row, column=2).value == expected

    def test_reinsurer_column_present(
        self, sufficiency_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(sufficiency_export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Combined Ratio")
        cell = ws.cell(row=row, column=3).value
        assert cell == pytest.approx(
            sufficiency_export.premium_sufficiency_reinsurer.combined_ratio
        )

    def test_cedant_only_no_reinsurer_column(self, tmp_path: Path) -> None:
        """Reinsurer column suppressed when no reinsurer result/sufficiency."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            premium_sufficiency_cedant=_make_sufficiency(),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Summary"]
        row = _find_row_with_label(ws, "Combined Ratio")
        assert ws.cell(row=row, column=3).value is None


# ---------------------------------------------------------------------------
# IFRS 17 Movement sheet (ADR-096, Epic 2 Slice 3b)
# ---------------------------------------------------------------------------


def _make_ifrs17_gross_cashflow(run_id: str, n_per: int = 36) -> CashFlowResult:
    """Synthetic GROSS CashFlowResult on a common grid for cohort aggregation."""
    premiums = np.full(n_per, 100.0, dtype=np.float64)
    claims = np.full(n_per, 10.0, dtype=np.float64)
    expenses = np.full(n_per, 5.0, dtype=np.float64)
    lapses = np.full(n_per, 2.0, dtype=np.float64)
    reserves = np.zeros(n_per, dtype=np.float64)
    return CashFlowResult(
        run_id=run_id,
        valuation_date=date(2025, 1, 1),
        basis="GROSS",
        assumption_set_version="v1",
        product_type="TERM",
        projection_months=n_per,
        time_index=np.arange(
            np.datetime64("2025-01"),
            np.datetime64("2025-01") + n_per,
            dtype="datetime64[M]",
        ),
        gross_premiums=premiums,
        death_claims=claims,
        lapse_surrenders=lapses,
        expenses=expenses,
        reserve_balance=reserves,
        reserve_increase=reserves.copy(),
        net_cash_flow=premiums - claims - lapses - expenses,
    )


def _make_movement_export() -> IFRS17MovementExport:
    """Two distinct issue-year cohorts (2022 @ 4%, 2024 @ 6%) valued 2025-01-01."""
    manager = IFRS17CohortManager(
        [
            IFRS17ContractInput(
                cashflows=_make_ifrs17_gross_cashflow("c2022"),
                issue_date=date(2022, 6, 1),
                locked_in_rate=0.04,
                ra_factor=0.05,
            ),
            IFRS17ContractInput(
                cashflows=_make_ifrs17_gross_cashflow("c2024"),
                issue_date=date(2024, 6, 1),
                locked_in_rate=0.06,
                ra_factor=0.05,
            ),
        ]
    )
    return IFRS17MovementExport(
        aggregate=manager.aggregate_movement_table(),
        cohorts=manager.cohort_movement_tables(),
    )


@pytest.fixture
def movement_export() -> DealPricingExport:
    """Minimal export augmented with the IFRS 17 movement tables."""
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_profit_result(),
        reinsurer_result=None,
        net_cashflows=_make_cashflows("NET"),
        ifrs17_movement=_make_movement_export(),
    )


class TestIFRS17MovementSheet:
    def test_sheet_omitted_when_none(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        assert "IFRS 17 Movement" not in load_workbook(out).sheetnames

    def test_sheet_present_and_appended_last(
        self, movement_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(movement_export, out)
        names = load_workbook(out).sheetnames
        assert "IFRS 17 Movement" in names
        # Appended last so all other sheet positions are unchanged.
        assert names[-1] == "IFRS 17 Movement"

    def test_title_and_aggregate_block(
        self, movement_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(movement_export, out)
        ws = load_workbook(out)["IFRS 17 Movement"]
        assert ws.cell(row=1, column=1).value == "IFRS 17 Analysis of Change (Movement)"
        # Aggregate block label present.
        _find_row_with_label(ws, "Aggregate (all cohorts)")

    def test_cohort_block_titles_carry_year_and_rate(
        self, movement_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(movement_export, out)
        ws = load_workbook(out)["IFRS 17 Movement"]
        # Ordered by issue year; each carries its locked-in rate.
        _find_row_with_label(ws, "Cohort 2022 — locked-in 4.00%")
        _find_row_with_label(ws, "Cohort 2024 — locked-in 6.00%")

    def test_component_labels_present(
        self, movement_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(movement_export, out)
        ws = load_workbook(out)["IFRS 17 Movement"]
        for label in (
            "Best Estimate Liability (BEL)",
            "Risk Adjustment (RA)",
            "Contractual Service Margin (CSM)",
            "Total Insurance Liability",
        ):
            _find_row_with_label(ws, label)

    def test_every_rendered_row_foots(
        self, movement_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """The disclosure's defining property — Opening + Σ movements == Closing —
        holds on every rendered data row (BEL / RA / CSM / total, aggregate and
        every cohort)."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(movement_export, out)
        ws = load_workbook(out)["IFRS 17 Movement"]
        checked = 0
        for r in range(1, ws.max_row + 1):
            year = ws.cell(row=r, column=1).value
            vals = [ws.cell(row=r, column=c).value for c in range(2, 7)]
            if isinstance(year, int) and all(isinstance(v, (int, float)) for v in vals):
                opening, new_business, interest, release, closing = vals
                np.testing.assert_allclose(
                    opening + new_business + interest + release, closing, atol=1e-6
                )
                checked += 1
        # 3 periods x 4 components x (1 aggregate + 2 cohorts) = 36 data rows.
        assert checked == 36

    def test_year_axis_is_one_based(
        self, movement_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(movement_export, out)
        ws = load_workbook(out)["IFRS 17 Movement"]
        # The first data row under the first "Year" header is reporting Year 1.
        header_row = _find_row_with_label(ws, "Year")
        assert ws.cell(row=header_row + 1, column=1).value == 1

    def test_workbook_roundtrips(self, movement_export: DealPricingExport, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(movement_export, out)
        wb = load_workbook(out)
        assert wb["IFRS 17 Movement"].max_row > 0


# ---------------------------------------------------------------------------
# ALM Duration Gap sheet (ADR-115, Asset/ALM Slice 4b-3)
# ---------------------------------------------------------------------------


def _make_duration_gap_side(
    *,
    asset_mv: float = 1_000_000.0,
    asset_mac: float = 10.0,
    asset_mod: float = 9.43,
    liab_pv: float = 800_000.0,
    liab_mac: float = 7.0,
    liab_mod: float = 6.6,
    valuation_yield: float = 0.06,
) -> DurationGapResult:
    """Build a self-consistent DurationGapResult for sheet-rendering tests.

    The duration / dollar-duration gap fields are derived from the asset/
    liability inputs so the rendered sheet values match a single source of
    truth (the writer only renders, it does not recompute).
    """
    dollar_asset = asset_mod * asset_mv
    dollar_liab = liab_mod * liab_pv
    return DurationGapResult(
        valuation_yield=valuation_yield,
        asset_market_value=asset_mv,
        asset_macaulay_duration=asset_mac,
        asset_modified_duration=asset_mod,
        liability_present_value=liab_pv,
        liability_macaulay_duration=liab_mac,
        liability_modified_duration=liab_mod,
        duration_gap=asset_mod - liab_mod,
        dollar_duration_asset=dollar_asset,
        dollar_duration_liability=dollar_liab,
        dollar_duration_gap=dollar_asset - dollar_liab,
    )


def _export_with_alm_gap(gap: DualDurationGap | None) -> DealPricingExport:
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_profit_result(),
        reinsurer_result=None,
        net_cashflows=_make_cashflows("NET"),
        alm_duration_gap=gap,
    )


class TestAlmDurationGapSheet:
    def test_sheet_absent_when_gap_none(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        """No asset portfolio → no ALM sheet (byte-identical to pre-ADR-115)."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        assert "ALM Duration Gap" not in load_workbook(out).sheetnames

    def test_sheet_absent_when_both_sides_none(self, tmp_path: Path) -> None:
        """An empty dual gap (both sides None) suppresses the sheet entirely."""
        out = tmp_path / "deal.xlsx"
        export = _export_with_alm_gap(DualDurationGap(reinsurer=None, cedant=None))
        write_deal_pricing_excel(export, out)
        assert "ALM Duration Gap" not in load_workbook(out).sheetnames

    def test_sheet_present_when_gap_supplied(self, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        gap = DualDurationGap(reinsurer=None, cedant=_make_duration_gap_side())
        write_deal_pricing_excel(_export_with_alm_gap(gap), out)
        assert "ALM Duration Gap" in load_workbook(out).sheetnames

    def test_cedant_only_omits_reinsurer_block(self, tmp_path: Path) -> None:
        """The YRT path: ceded reserve ~0 → reinsurer side None, only cedant block."""
        out = tmp_path / "deal.xlsx"
        gap = DualDurationGap(reinsurer=None, cedant=_make_duration_gap_side())
        write_deal_pricing_excel(_export_with_alm_gap(gap), out)
        ws = load_workbook(out)["ALM Duration Gap"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Cedant (retained)" in labels
        assert "Reinsurer (ceded)" not in labels

    def test_both_sides_render_reinsurer_first(self, tmp_path: Path) -> None:
        """Coinsurance path: both sides defined, reinsurer (headline) rendered first."""
        out = tmp_path / "deal.xlsx"
        gap = DualDurationGap(
            reinsurer=_make_duration_gap_side(liab_pv=900_000.0),
            cedant=_make_duration_gap_side(liab_pv=100_000.0),
        )
        write_deal_pricing_excel(_export_with_alm_gap(gap), out)
        ws = load_workbook(out)["ALM Duration Gap"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        rei_row = labels.index("Reinsurer (ceded)")
        ced_row = labels.index("Cedant (retained)")
        assert rei_row < ced_row

    def test_asset_and_liability_values_match_model(self, tmp_path: Path) -> None:
        """Rendered Asset/Liability cells equal the DurationGapResult fields."""
        out = tmp_path / "deal.xlsx"
        side = _make_duration_gap_side()
        gap = DualDurationGap(reinsurer=None, cedant=side)
        write_deal_pricing_excel(_export_with_alm_gap(gap), out)
        ws = load_workbook(out)["ALM Duration Gap"]
        row = _find_row_with_label(ws, "Value ($)")
        assert ws.cell(row=row, column=2).value == pytest.approx(side.asset_market_value)
        assert ws.cell(row=row, column=3).value == pytest.approx(side.liability_present_value)
        mod_row = _find_row_with_label(ws, "Modified duration (yrs)")
        assert ws.cell(row=mod_row, column=2).value == pytest.approx(side.asset_modified_duration)
        assert ws.cell(row=mod_row, column=3).value == pytest.approx(
            side.liability_modified_duration
        )

    def test_net_gap_rows_match_model(self, tmp_path: Path) -> None:
        """The net gap rows (yield / duration gap / dollar gap) match the model."""
        out = tmp_path / "deal.xlsx"
        side = _make_duration_gap_side()
        gap = DualDurationGap(reinsurer=None, cedant=side)
        write_deal_pricing_excel(_export_with_alm_gap(gap), out)
        ws = load_workbook(out)["ALM Duration Gap"]
        y_row = _find_row_with_label(ws, "Valuation yield")
        assert ws.cell(row=y_row, column=2).value == pytest.approx(side.valuation_yield)
        g_row = _find_row_with_label(ws, "Duration gap (yrs)")
        assert ws.cell(row=g_row, column=2).value == pytest.approx(side.duration_gap)
        dd_row = _find_row_with_label(ws, "Dollar-duration gap ($·yr)")
        assert ws.cell(row=dd_row, column=2).value == pytest.approx(side.dollar_duration_gap)

    def test_sheet_appended_after_other_sheets(self, tmp_path: Path) -> None:
        """ALM sheet is appended last, leaving the existing sheet order intact."""
        out = tmp_path / "deal.xlsx"
        gap = DualDurationGap(reinsurer=None, cedant=_make_duration_gap_side())
        write_deal_pricing_excel(_export_with_alm_gap(gap), out)
        names = load_workbook(out).sheetnames
        assert names[-1] == "ALM Duration Gap"
        assert names[:3] == ["Summary", "Cash Flows", "Assumptions"]

    def test_workbook_roundtrips(self, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        gap = DualDurationGap(reinsurer=None, cedant=_make_duration_gap_side())
        write_deal_pricing_excel(_export_with_alm_gap(gap), out)
        wb = load_workbook(out)
        assert wb["ALM Duration Gap"].max_row > 0


# ---------------------------------------------------------------------------
# Treaty-terms panel on the Assumptions sheet (ADR-124)
# ---------------------------------------------------------------------------


def _make_expense_allowance(*, sliding: bool = False):
    """A sample ExpenseAllowance; optionally with a monotone sliding scale."""
    from polaris_re.reinsurance.expense_allowance import (
        ExpenseAllowance,
        ExpenseAllowanceBand,
    )

    scale = None
    if sliding:
        scale = [
            ExpenseAllowanceBand(max_loss_ratio=0.60, allowance_pct=0.30),
            ExpenseAllowanceBand(max_loss_ratio=0.80, allowance_pct=0.20),
            ExpenseAllowanceBand(max_loss_ratio=1.00, allowance_pct=0.10),
        ]
    return ExpenseAllowance(
        first_year_pct=0.55,
        renewal_pct=0.12,
        sliding_scale=scale,
    )


def _make_experience_refund():
    from polaris_re.reinsurance.experience_refund import ExperienceRefund

    return ExperienceRefund(
        refund_pct=0.50,
        retention=25_000.0,
        reinsurer_margin_pct=0.05,
        interest_rate=0.04,
    )


def _export_with_terms(*, allowance=None, refund=None) -> DealPricingExport:
    return DealPricingExport(
        deal_meta=_make_deal_meta(),
        assumptions_meta=_make_assumptions_meta(),
        cedant_result=_make_profit_result(),
        reinsurer_result=None,
        net_cashflows=_make_cashflows("NET"),
        expense_allowance=allowance,
        experience_refund=refund,
    )


class TestTreatyTermsPanel:
    """Slice 3b-2b-2 of the expense-allowance epic (ADR-124).

    The Assumptions sheet picks up an optional "Treaty Terms" panel showing the
    sliding-scale expense allowance and/or experience refund the deal was priced
    with. Suppressed when neither term is supplied, so every workbook priced
    without these terms (the common path) stays byte-identical to pre-ADR-124
    output.
    """

    def test_panel_absent_by_default(
        self, minimal_export: DealPricingExport, tmp_path: Path
    ) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(minimal_export, out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Treaty Terms" not in labels
        assert "Expense Allowance" not in labels
        assert "Experience Refund" not in labels

    def test_allowance_rows_rendered(self, tmp_path: Path) -> None:
        allowance = _make_expense_allowance()
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(_export_with_terms(allowance=allowance), out)
        ws = load_workbook(out)["Assumptions"]
        assert "Treaty Terms" in [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        fy_row = _find_row_with_label(ws, "First-Year Allowance %")
        assert ws.cell(row=fy_row, column=2).value == pytest.approx(0.55)
        rn_row = _find_row_with_label(ws, "Renewal Allowance %")
        assert ws.cell(row=rn_row, column=2).value == pytest.approx(0.12)

    def test_sliding_scale_bands_rendered(self, tmp_path: Path) -> None:
        allowance = _make_expense_allowance(sliding=True)
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(_export_with_terms(allowance=allowance), out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Sliding Scale (renewal)" in labels
        # One row per band, with the band's allowance % in column B.
        band_row = _find_row_with_label(ws, "≤ loss ratio 0.8")
        assert ws.cell(row=band_row, column=2).value == pytest.approx(0.20)

    def test_no_sliding_scale_rows_when_flat(self, tmp_path: Path) -> None:
        allowance = _make_expense_allowance(sliding=False)
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(_export_with_terms(allowance=allowance), out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Sliding Scale (renewal)" not in labels

    def test_refund_rows_rendered(self, tmp_path: Path) -> None:
        refund = _make_experience_refund()
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(_export_with_terms(refund=refund), out)
        ws = load_workbook(out)["Assumptions"]
        assert "Experience Refund" in [
            ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)
        ]
        assert ws.cell(row=_find_row_with_label(ws, "Refund %"), column=2).value == pytest.approx(
            0.50
        )
        assert ws.cell(row=_find_row_with_label(ws, "Retention"), column=2).value == pytest.approx(
            25_000.0
        )
        assert ws.cell(
            row=_find_row_with_label(ws, "Reinsurer Margin %"), column=2
        ).value == pytest.approx(0.05)
        assert ws.cell(
            row=_find_row_with_label(ws, "Interest Rate"), column=2
        ).value == pytest.approx(0.04)

    def test_allowance_only_omits_refund_section(self, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(_export_with_terms(allowance=_make_expense_allowance()), out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Expense Allowance" in labels
        assert "Experience Refund" not in labels

    def test_refund_only_omits_allowance_section(self, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(_export_with_terms(refund=_make_experience_refund()), out)
        ws = load_workbook(out)["Assumptions"]
        labels = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
        assert "Experience Refund" in labels
        assert "Expense Allowance" not in labels

    def test_both_sections_present_and_ordered(self, tmp_path: Path) -> None:
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(
            _export_with_terms(
                allowance=_make_expense_allowance(sliding=True),
                refund=_make_experience_refund(),
            ),
            out,
        )
        ws = load_workbook(out)["Assumptions"]
        allow_row = _find_row_with_label(ws, "Expense Allowance")
        refund_row = _find_row_with_label(ws, "Experience Refund")
        # Allowance section precedes the refund section.
        assert allow_row < refund_row

    def test_panel_coexists_with_rated_block(self, tmp_path: Path) -> None:
        """Both panels render on the Assumptions sheet without overlapping."""
        export = DealPricingExport(
            deal_meta=_make_deal_meta(),
            assumptions_meta=_make_assumptions_meta(),
            cedant_result=_make_profit_result(),
            reinsurer_result=None,
            net_cashflows=_make_cashflows("NET"),
            rated_block=_make_rated_block(),
            expense_allowance=_make_expense_allowance(),
        )
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(export, out)
        ws = load_workbook(out)["Assumptions"]
        rated_row = _find_row_with_label(ws, "Rated Block")
        terms_row = _find_row_with_label(ws, "Treaty Terms")
        # Treaty-terms panel is appended after the rated-block panel, and the
        # rated-block metrics are not clobbered by it.
        assert terms_row > rated_row
        assert (
            ws.cell(row=_find_row_with_label(ws, "Policies Rated"), column=2).value
            == _make_rated_block().n_rated
        )

    def test_sheet_order_unchanged(self, tmp_path: Path) -> None:
        """The panel lives inside Assumptions; it adds no sheet."""
        out = tmp_path / "deal.xlsx"
        write_deal_pricing_excel(
            _export_with_terms(
                allowance=_make_expense_allowance(),
                refund=_make_experience_refund(),
            ),
            out,
        )
        assert load_workbook(out).sheetnames == ["Summary", "Cash Flows", "Assumptions"]
