"""
Excel output formatters.

Two workbooks are produced from this module:

* ``write_rate_schedule_excel`` — YRT rate schedule, one sheet per
  sex/smoker combination, used by ``polaris rate-schedule`` (pre-existing).
* ``write_deal_pricing_excel`` — deal-pricing committee packet, covering
  Summary / Cash Flows / Assumptions / Sensitivity, plus optional
  Gross / Ceded basis sheets (ADR-080) and a combined Gross / Ceded / Net
  comparison sheet (ADR-081). See ADR-045 for the workbook schema and the
  rationale for the fixed sheets.

The deal-pricing export takes a ``DealPricingExport`` dataclass bundling
the ``ProfitTestResult`` objects, the ``CashFlowResult`` used by the
profit test, and structured metadata (``DealMetaExport`` /
``AssumptionsMetaExport`` / ``ScenarioMetric``). This keeps the writer
signature stable while the CLI wiring (Slice 2) grows above it.
"""

from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import TYPE_CHECKING

import numpy as np
import polars as pl

from polaris_re.analytics.premium_sufficiency import PremiumSufficiencyResult
from polaris_re.analytics.profit_test import ProfitResultWithCapital, ProfitTestResult
from polaris_re.core.cashflow import CashFlowResult

if TYPE_CHECKING:
    # openpyxl ships via the `[tables]` extra; type-only imports here keep the
    # module importable without the extra installed.
    from openpyxl.workbook.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

    from polaris_re.reinsurance.yrt_rate_table import YRTRateTable

__all__ = [
    "AssumptionsMetaExport",
    "DealMetaExport",
    "DealPricingExport",
    "RatedBlockExport",
    "ScenarioMetric",
    "write_deal_pricing_excel",
    "write_rate_schedule_excel",
    "write_yrt_rate_table_excel",
]


# ---------------------------------------------------------------------------
# Deal-pricing export DTOs
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class DealMetaExport:
    """Deal-level metadata rendered on the Assumptions sheet."""

    product_type: str
    n_policies: int
    face_amount: float
    treaty_type: str | None
    cession_pct: float | None
    hurdle_rate: float
    discount_rate: float
    projection_years: int
    valuation_date: date


@dataclass(frozen=True)
class AssumptionsMetaExport:
    """Assumption-set metadata rendered on the Assumptions sheet."""

    mortality_source: str
    mortality_multiplier: float
    lapse_description: str
    assumption_set_version: str


@dataclass(frozen=True)
class ScenarioMetric:
    """One row on the Sensitivity sheet."""

    name: str
    pv_profits: float
    irr: float | None
    profit_margin: float | None


@dataclass(frozen=True)
class RatedBlockExport:
    """Block-level substandard-rating composition (ADR-068).

    Mirrors the dict returned by ``polaris_re.utils.rating.rating_composition``
    in typed form so the deal-pricing workbook's Assumptions sheet can
    render the same numbers the CLI Rich panel shows. The writer suppresses
    the panel when ``n_rated == 0``, so all-standard blocks stay
    byte-identical to pre-ADR-068 output.
    """

    n_policies: int
    n_rated: int
    pct_rated_by_count: float
    pct_rated_by_face: float
    face_weighted_mean_multiplier: float
    max_multiplier: float
    max_flat_extra_per_1000: float


@dataclass(frozen=True)
class DealPricingExport:
    """Bundle of everything a deal-pricing workbook needs.

    ``scenario_results=None`` suppresses the Sensitivity sheet. A
    ``reinsurer_result=None`` suppresses the reinsurer column on the
    Summary sheet — both signal "ceded side does not apply to this deal".
    ``gross_cashflows=None`` / ``ceded_cashflows=None`` suppress the
    ``Gross Cash Flows`` / ``Ceded Cash Flows`` sheets respectively
    (ADR-080); when populated they render the same annual-rollup layout as
    the NET ``Cash Flows`` sheet. When BOTH are populated, a combined
    ``Cash Flow Comparison`` sheet (ADR-081) is also written, placing the
    three bases' per-year Net Cash Flow side by side with a ``Gross - Ceded``
    check column, followed by a ``Line Item Comparison`` sheet (ADR-086) that
    breaks the same three-basis comparison out per component line item
    (premiums, claims, surrenders, expenses, reserve increase).
    ``yrt_rate_table=None`` suppresses the ``YRT Rate Table`` sheet
    (only added by ADR-052 when the deal was priced with a tabular
    schedule). ``rated_block=None`` (or ``n_rated == 0``) suppresses the
    rated-block panel appended to the Assumptions sheet (ADR-068).
    """

    deal_meta: DealMetaExport
    assumptions_meta: AssumptionsMetaExport
    cedant_result: ProfitTestResult
    reinsurer_result: ProfitTestResult | None
    net_cashflows: CashFlowResult
    gross_cashflows: CashFlowResult | None = None
    ceded_cashflows: CashFlowResult | None = None
    scenario_results: list[ScenarioMetric] | None = None
    yrt_rate_table: "YRTRateTable | None" = None
    rated_block: RatedBlockExport | None = None
    # Premium-sufficiency panel (ADR-083). When ``premium_sufficiency_cedant``
    # is populated, a "Premium Sufficiency" block is appended to the Summary
    # sheet (cedant column always; reinsurer column when
    # ``premium_sufficiency_reinsurer`` is also populated). ``None`` suppresses
    # the panel entirely, keeping pre-ADR-083 workbooks byte-identical.
    premium_sufficiency_cedant: PremiumSufficiencyResult | None = None
    premium_sufficiency_reinsurer: PremiumSufficiencyResult | None = None


# ---------------------------------------------------------------------------
# Rate-schedule writer (unchanged)
# ---------------------------------------------------------------------------


def write_rate_schedule_excel(df: pl.DataFrame, path: Path) -> None:
    """
    Write a YRT rate schedule DataFrame to a formatted Excel workbook.

    Creates one sheet per sex/smoker combination. Each sheet has issue_age
    as rows and rate_per_1000 as the value column.

    Args:
        df:   Rate schedule DataFrame from ``YRTRateSchedule.generate()``.
              Expected columns: issue_age, sex, smoker_status, policy_term,
              rate_per_1000, irr.
        path: Output .xlsx file path.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise ImportError(
            "openpyxl required for Excel output. Install via: uv sync --extra tables"
        ) from exc

    wb = Workbook()
    # Remove default sheet
    if wb.active is not None:
        wb.remove(wb.active)

    # Group by sex + smoker_status
    groups = df.group_by(["sex", "smoker_status"]).agg(pl.all())

    for row in groups.iter_rows(named=True):
        sex = row["sex"]
        smoker = row["smoker_status"]
        sheet_name = f"{sex}_{smoker}"

        ws = wb.create_sheet(title=sheet_name)

        # Header row
        headers = ["Issue Age", "Policy Term", "Rate per $1,000", "IRR"]
        header_font = Font(bold=True)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Get the ages and rates for this group
        ages = row["issue_age"]
        terms = row["policy_term"]
        rates = row["rate_per_1000"]
        irrs = row["irr"]

        for i in range(len(ages)):
            row_idx = i + 2
            ws.cell(row=row_idx, column=1, value=ages[i])
            ws.cell(row=row_idx, column=2, value=terms[i])

            rate_cell = ws.cell(row=row_idx, column=3, value=rates[i])
            rate_cell.number_format = "0.0000"

            irr_cell = ws.cell(row=row_idx, column=4, value=irrs[i])
            irr_cell.number_format = "0.00%"

        # Auto-width columns
        for col_idx in range(1, 5):
            ws.column_dimensions[chr(64 + col_idx)].width = 16

    # Add summary sheet
    ws_summary = wb.create_sheet(title="Summary", index=0)
    ws_summary.cell(row=1, column=1, value="YRT Rate Schedule Summary").font = Font(
        bold=True, size=14
    )
    ws_summary.cell(row=3, column=1, value="Generated by Polaris RE")
    ws_summary.cell(row=4, column=1, value=f"Total rate cells: {len(df)}")

    if "policy_term" in df.columns:
        terms = df["policy_term"].unique().to_list()
        ws_summary.cell(row=5, column=1, value=f"Policy terms: {terms}")

    wb.save(path)


# ---------------------------------------------------------------------------
# Deal-pricing writer (ADR-045)
# ---------------------------------------------------------------------------


# Canonical column order for the Cash Flows sheet. Kept as a module-level
# constant so tests and docs reference the same list.
_CASH_FLOW_COLUMNS: tuple[str, ...] = (
    "Year",
    "Gross Premiums",
    "Death Claims",
    "Lapse Surrenders",
    "Expenses",
    "Reserve Increase",
    "Net Cash Flow",
)

# Columns for the combined Gross / Ceded / Net comparison sheet (ADR-081).
# Each value is that basis' per-year Net Cash Flow; "Gross - Ceded" is a
# visual check column that equals the "Net" column by construction
# (Net = Gross - Ceded — the treaty decomposition).
_CASH_FLOW_COMPARISON_COLUMNS: tuple[str, ...] = (
    "Year",
    "Gross",
    "Ceded",
    "Net",
    "Gross - Ceded",
)

# Component cash-flow line items broken out on the per-line-item comparison
# sheet (ADR-086): every basis-sheet column except "Year" and the bottom-line
# "Net Cash Flow" (which the Cash Flow Comparison sheet already diffs). Kept as
# a module-level constant so tests and docs reference the same list; the order
# mirrors ``_CASH_FLOW_COLUMNS``.
_LINE_ITEM_COMPARISON_LINE_ITEMS: tuple[str, ...] = (
    "Gross Premiums",
    "Death Claims",
    "Lapse Surrenders",
    "Expenses",
    "Reserve Increase",
)

# Header row for the per-line-item comparison sheet: "Year" followed by a
# (Gross, Ceded, Net) triplet per line item, so each component's three bases
# sit side by side.
_LINE_ITEM_COMPARISON_COLUMNS: tuple[str, ...] = (
    "Year",
    *(
        f"{item} ({basis})"
        for item in _LINE_ITEM_COMPARISON_LINE_ITEMS
        for basis in ("Gross", "Ceded", "Net")
    ),
)

_SUMMARY_METRICS: tuple[str, ...] = (
    "Hurdle Rate",
    "PV Profits",
    "PV Premiums",
    "Profit Margin",
    "IRR",
    "Breakeven Year",
    "Total Undiscounted Profit",
)

# LICAT-capital rows appended to the Summary sheet when at least one of
# the rendered profit-test results carries capital metrics (ADR-049).
# Order matches the CLI Rich table for readability.
_CAPITAL_METRICS: tuple[str, ...] = (
    "Peak Capital",
    "PV Capital (stock)",
    "PV Capital Strain",
    "Return on Capital",
    "Capital-Adjusted IRR",
)

# Premium-sufficiency rows appended to the Summary sheet when
# ``premium_sufficiency_cedant`` is populated (ADR-083). The discount rate
# here is the valuation rate used by the analyzer, not the profit hurdle.
# ``PV Claims`` / ``PV Surrenders`` break out ``PV Benefits`` into its two
# components so the combined ratio's numerator is readable line-by-line
# (ADR-084); they sum to ``PV Benefits`` by construction.
_SUFFICIENCY_METRICS: tuple[str, ...] = (
    "Sufficiency Discount Rate",
    "Sufficiency Target Margin",
    "PV Claims",
    "PV Surrenders",
    "PV Benefits",
    "PV Expenses",
    "Sufficiency Margin",
    "Loss Ratio",
    "Expense Ratio",
    "Combined Ratio",
    "Premium Sufficient",
)


def write_deal_pricing_excel(export: DealPricingExport, path: Path) -> None:
    """Write a formatted deal-pricing workbook (see ADR-045 / ADR-052).

    Sheets produced (in order):
        1. Summary           — cedant (and optional reinsurer) profit metrics.
        2. Gross Cash Flows  — OMITTED when ``export.gross_cashflows`` is None;
                               otherwise the annual rollup of the GROSS
                               CashFlowResult (ADR-080).
        3. Ceded Cash Flows  — OMITTED when ``export.ceded_cashflows`` is None;
                               otherwise the annual rollup of the CEDED
                               CashFlowResult (ADR-080).
        4. Cash Flows        — annual rollup of the NET CashFlowResult.
        5. Cash Flow Comparison — OMITTED unless BOTH ``export.gross_cashflows``
                               and ``export.ceded_cashflows`` are populated;
                               otherwise the per-year Net Cash Flow of all
                               three bases side by side with a ``Gross - Ceded``
                               check column (ADR-081).
        6. Line Item Comparison — OMITTED under the same gate as sheet 5;
                               otherwise each component line item (premiums,
                               claims, surrenders, expenses, reserve increase)
                               of all three bases side by side, so the ceded
                               share is visible per line, not just on the net
                               total (ADR-086).
        7. Assumptions       — deal + assumption-set metadata.
        8. Sensitivity       — OMITTED when ``export.scenario_results`` is None.
        9. YRT Rate Table    — OMITTED when ``export.yrt_rate_table`` is None;
                               otherwise one block per (sex, smoker) cohort
                               (ADR-052).

    Args:
        export: Bundle of all data required for the sheets.
        path:   Output .xlsx file path. Parent directory must exist.

    Raises:
        ImportError: If openpyxl is not installed.
    """
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl required for Excel output. Install via: uv sync --extra tables"
        ) from exc

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    _write_summary_sheet(wb, export)
    # Gross / Ceded / Net cash-flow block (ADR-080). The Gross and Ceded
    # sheets are written only when their optional DTO fields are populated,
    # so a net-only export remains byte-identical (the NET sheet keeps its
    # canonical "Cash Flows" title and immediately follows Summary).
    if export.gross_cashflows is not None:
        _write_cash_flows_sheet(wb, export.gross_cashflows, "Gross Cash Flows")
    if export.ceded_cashflows is not None:
        _write_cash_flows_sheet(wb, export.ceded_cashflows, "Ceded Cash Flows")
    _write_cash_flows_sheet(wb, export.net_cashflows, "Cash Flows")
    # Combined Gross / Ceded / Net comparison sheet (ADR-081). Written only
    # when both ceded-side bases are present, so net-only and gross-only
    # exports stay byte-identical (the comparison needs all three bases).
    if export.gross_cashflows is not None and export.ceded_cashflows is not None:
        _write_cash_flow_comparison_sheet(
            wb, export.gross_cashflows, export.ceded_cashflows, export.net_cashflows
        )
        # Per-line-item Gross / Ceded / Net comparison (ADR-086). Same gate as
        # the bottom-line comparison sheet above: written only when all three
        # bases are present, so net-only / gross-only exports stay byte-identical.
        _write_line_item_comparison_sheet(
            wb, export.gross_cashflows, export.ceded_cashflows, export.net_cashflows
        )
    _write_assumptions_sheet(wb, export)
    if export.scenario_results is not None:
        _write_sensitivity_sheet(wb, export.scenario_results)
    if export.yrt_rate_table is not None:
        _write_yrt_rate_table_sheet(wb, export.yrt_rate_table)

    wb.save(path)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------


def _fmt_rate(value: float | None) -> object:
    """Return a float (for Excel % formatting) or 'N/A' when suppressed."""
    return "N/A" if value is None else float(value)


def _fmt_int(value: int | None) -> object:
    return "N/A" if value is None else int(value)


def _write_summary_sheet(wb: "Workbook", export: DealPricingExport) -> None:
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title="Summary")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)
    centre = Alignment(horizontal="center")

    ws.cell(row=1, column=1, value="Deal Pricing Summary").font = title_font

    # Header row at row 3: Metric | Cedant | [Reinsurer]
    ws.cell(row=3, column=1, value="Metric").font = header_font
    ws.cell(row=3, column=2, value="Cedant (NET)").font = header_font
    ws.cell(row=3, column=2).alignment = centre
    if export.reinsurer_result is not None:
        ws.cell(row=3, column=3, value="Reinsurer").font = header_font
        ws.cell(row=3, column=3).alignment = centre

    for i, metric in enumerate(_SUMMARY_METRICS):
        row_idx = 4 + i
        ws.cell(row=row_idx, column=1, value=metric).font = header_font
        _write_metric_cell(ws, row_idx, 2, metric, export.cedant_result)
        if export.reinsurer_result is not None:
            _write_metric_cell(ws, row_idx, 3, metric, export.reinsurer_result)

    # Optional LICAT capital block — appended only when at least one
    # rendered result is a ProfitResultWithCapital (ADR-049). Workbooks
    # produced without --capital remain byte-identical pre-Slice-3.
    capital_present = isinstance(export.cedant_result, ProfitResultWithCapital) or (
        export.reinsurer_result is not None
        and isinstance(export.reinsurer_result, ProfitResultWithCapital)
    )
    next_row = 4 + len(_SUMMARY_METRICS)
    if capital_present:
        for i, metric in enumerate(_CAPITAL_METRICS):
            row_idx = next_row + i
            ws.cell(row=row_idx, column=1, value=metric).font = header_font
            _write_capital_cell(ws, row_idx, 2, metric, export.cedant_result)
            if export.reinsurer_result is not None:
                _write_capital_cell(ws, row_idx, 3, metric, export.reinsurer_result)
        next_row += len(_CAPITAL_METRICS)

    # Optional premium-sufficiency block — appended only when populated
    # (ADR-083). Workbooks produced without sufficiency data remain
    # byte-identical to pre-ADR-083 output. The reinsurer column is written
    # only when both a reinsurer profit result and a reinsurer sufficiency
    # result are present.
    if export.premium_sufficiency_cedant is not None:
        for i, metric in enumerate(_SUFFICIENCY_METRICS):
            row_idx = next_row + i
            ws.cell(row=row_idx, column=1, value=metric).font = header_font
            _write_sufficiency_cell(ws, row_idx, 2, metric, export.premium_sufficiency_cedant)
            if (
                export.reinsurer_result is not None
                and export.premium_sufficiency_reinsurer is not None
            ):
                _write_sufficiency_cell(
                    ws, row_idx, 3, metric, export.premium_sufficiency_reinsurer
                )

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20
    if export.reinsurer_result is not None:
        ws.column_dimensions["C"].width = 20


def _write_metric_cell(
    ws: "Worksheet",
    row: int,
    col: int,
    metric: str,
    result: ProfitTestResult,
) -> None:
    """Write one Summary-sheet metric cell with appropriate number format."""
    cell = ws.cell(row=row, column=col)
    if metric == "Hurdle Rate":
        cell.value = result.hurdle_rate
        cell.number_format = "0.00%"
    elif metric == "PV Profits":
        cell.value = float(result.pv_profits)
        cell.number_format = "$#,##0"
    elif metric == "PV Premiums":
        cell.value = float(result.pv_premiums)
        cell.number_format = "$#,##0"
    elif metric == "Profit Margin":
        cell.value = _fmt_rate(result.profit_margin)
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    elif metric == "IRR":
        cell.value = _fmt_rate(result.irr)
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    elif metric == "Breakeven Year":
        cell.value = _fmt_int(result.breakeven_year)
    elif metric == "Total Undiscounted Profit":
        cell.value = float(result.total_undiscounted_profit)
        cell.number_format = "$#,##0"


def _write_capital_cell(
    ws: "Worksheet",
    row: int,
    col: int,
    metric: str,
    result: ProfitTestResult,
) -> None:
    """Write one Summary-sheet LICAT capital cell.

    Renders ``"N/A"`` when the result is not a
    ``ProfitResultWithCapital`` (e.g. mixed runs where one side has no
    treaty and therefore no reinsurer capital).
    """
    cell = ws.cell(row=row, column=col)
    if not isinstance(result, ProfitResultWithCapital):
        cell.value = "N/A"
        return
    if metric == "Peak Capital":
        cell.value = float(result.peak_capital)
        cell.number_format = "$#,##0"
    elif metric == "PV Capital (stock)":
        cell.value = float(result.pv_capital)
        cell.number_format = "$#,##0"
    elif metric == "PV Capital Strain":
        cell.value = float(result.pv_capital_strain)
        cell.number_format = "$#,##0"
    elif metric == "Return on Capital":
        cell.value = _fmt_rate(result.return_on_capital)
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    elif metric == "Capital-Adjusted IRR":
        cell.value = _fmt_rate(result.capital_adjusted_irr)
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"


def _write_sufficiency_cell(
    ws: "Worksheet",
    row: int,
    col: int,
    metric: str,
    result: PremiumSufficiencyResult,
) -> None:
    """Write one Summary-sheet premium-sufficiency cell (ADR-083)."""
    cell = ws.cell(row=row, column=col)
    if metric == "Sufficiency Discount Rate":
        cell.value = float(result.discount_rate)
        cell.number_format = "0.00%"
    elif metric == "Sufficiency Target Margin":
        cell.value = float(result.target_margin)
        cell.number_format = "0.00%"
    elif metric == "PV Claims":
        # Per-line-item breakdown of PV Benefits (ADR-084).
        cell.value = float(result.pv_claims)
        cell.number_format = "$#,##0"
    elif metric == "PV Surrenders":
        # Per-line-item breakdown of PV Benefits (ADR-084).
        cell.value = float(result.pv_surrenders)
        cell.number_format = "$#,##0"
    elif metric == "PV Benefits":
        cell.value = float(result.pv_benefits)
        cell.number_format = "$#,##0"
    elif metric == "PV Expenses":
        cell.value = float(result.pv_expenses)
        cell.number_format = "$#,##0"
    elif metric == "Sufficiency Margin":
        cell.value = float(result.sufficiency_margin)
        cell.number_format = "$#,##0"
    elif metric == "Loss Ratio":
        cell.value = _fmt_rate(result.loss_ratio)
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    elif metric == "Expense Ratio":
        cell.value = _fmt_rate(result.expense_ratio)
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    elif metric == "Combined Ratio":
        cell.value = _fmt_rate(result.combined_ratio)
        if isinstance(cell.value, float):
            cell.number_format = "0.00%"
    elif metric == "Premium Sufficient":
        cell.value = "Yes" if result.is_sufficient else "No"


def _write_cash_flows_sheet(wb: "Workbook", cashflows: CashFlowResult, title: str) -> None:
    """Render one annual cash-flow rollup sheet for the given basis.

    Used for the NET ("Cash Flows"), GROSS ("Gross Cash Flows") and CEDED
    ("Ceded Cash Flows") bases; the column layout (``_CASH_FLOW_COLUMNS``)
    is identical across all three so the sheets read consistently.
    """
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title=title)
    header_font = Font(bold=True)
    centre = Alignment(horizontal="center")

    for col_idx, header in enumerate(_CASH_FLOW_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = centre

    annual = _aggregate_monthly_to_annual(cashflows)
    for row_idx, row in enumerate(annual, start=2):
        ws.cell(row=row_idx, column=1, value=row["Year"])
        ws.cell(row=row_idx, column=2, value=row["Gross Premiums"]).number_format = "$#,##0"
        ws.cell(row=row_idx, column=3, value=row["Death Claims"]).number_format = "$#,##0"
        ws.cell(row=row_idx, column=4, value=row["Lapse Surrenders"]).number_format = "$#,##0"
        ws.cell(row=row_idx, column=5, value=row["Expenses"]).number_format = "$#,##0"
        ws.cell(row=row_idx, column=6, value=row["Reserve Increase"]).number_format = "$#,##0"
        ws.cell(row=row_idx, column=7, value=row["Net Cash Flow"]).number_format = "$#,##0"

    ws.column_dimensions["A"].width = 8
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 18


def _write_cash_flow_comparison_sheet(
    wb: "Workbook",
    gross: CashFlowResult,
    ceded: CashFlowResult,
    net: CashFlowResult,
) -> None:
    """Render the combined Gross / Ceded / Net comparison sheet (ADR-081).

    Places the per-year Net Cash Flow of all three bases side by side so a
    committee can read the waterfall (Net = Gross - Ceded) on one sheet
    instead of diffing across the three separate basis sheets. The trailing
    ``Gross - Ceded`` column is a visual check that equals the ``Net`` column
    by construction (``treaty.apply`` returns ``net = gross - ceded``).

    Each basis is rolled up to annual rows with the same
    ``_aggregate_monthly_to_annual`` helper the basis sheets use, so the
    Year axis and per-year values match those sheets exactly.
    """
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title="Cash Flow Comparison")
    header_font = Font(bold=True)
    centre = Alignment(horizontal="center")

    for col_idx, header in enumerate(_CASH_FLOW_COMPARISON_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = centre

    gross_annual = _aggregate_monthly_to_annual(gross)
    ceded_annual = _aggregate_monthly_to_annual(ceded)
    net_annual = _aggregate_monthly_to_annual(net)

    for row_idx, (g_row, c_row, n_row) in enumerate(
        zip(gross_annual, ceded_annual, net_annual, strict=True), start=2
    ):
        g = g_row["Net Cash Flow"]
        c = c_row["Net Cash Flow"]
        n = n_row["Net Cash Flow"]
        ws.cell(row=row_idx, column=1, value=n_row["Year"])
        ws.cell(row=row_idx, column=2, value=g).number_format = "$#,##0"
        ws.cell(row=row_idx, column=3, value=c).number_format = "$#,##0"
        ws.cell(row=row_idx, column=4, value=n).number_format = "$#,##0"
        ws.cell(row=row_idx, column=5, value=g - c).number_format = "$#,##0"

    ws.column_dimensions["A"].width = 8
    for col in "BCDE":
        ws.column_dimensions[col].width = 18


def _write_line_item_comparison_sheet(
    wb: "Workbook",
    gross: CashFlowResult,
    ceded: CashFlowResult,
    net: CashFlowResult,
) -> None:
    """Render the per-line-item Gross / Ceded / Net comparison sheet (ADR-086).

    The ``Cash Flow Comparison`` sheet (ADR-081) diffs only the bottom-line Net
    Cash Flow across the three bases. This sheet extends that side-by-side
    treatment to every *component* line item (premiums, claims, surrenders,
    expenses, reserve increase), so a committee can see where the ceded share
    concentrates rather than only the net result. For each line item the three
    bases sit in a (Gross, Ceded, Net) triplet; the Net column equals
    Gross - Ceded component-by-component (``treaty.apply`` returns
    ``net = gross - ceded`` for every cash-flow line, not just the total).

    Each basis is rolled up to annual rows with the same
    ``_aggregate_monthly_to_annual`` helper the basis sheets use, so the Year
    axis and per-year values match those sheets exactly.
    """
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title="Line Item Comparison")
    header_font = Font(bold=True)
    centre = Alignment(horizontal="center")

    for col_idx, header in enumerate(_LINE_ITEM_COMPARISON_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = centre

    gross_annual = _aggregate_monthly_to_annual(gross)
    ceded_annual = _aggregate_monthly_to_annual(ceded)
    net_annual = _aggregate_monthly_to_annual(net)

    for row_idx, (g_row, c_row, n_row) in enumerate(
        zip(gross_annual, ceded_annual, net_annual, strict=True), start=2
    ):
        ws.cell(row=row_idx, column=1, value=n_row["Year"])
        col = 2
        for item in _LINE_ITEM_COMPARISON_LINE_ITEMS:
            for basis_row in (g_row, c_row, n_row):
                ws.cell(row=row_idx, column=col, value=basis_row[item]).number_format = "$#,##0"
                col += 1

    ws.column_dimensions["A"].width = 8
    for col_idx in range(2, len(_LINE_ITEM_COMPARISON_COLUMNS) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18


def _aggregate_monthly_to_annual(cf: CashFlowResult) -> list[dict[str, float]]:
    """Collapse a monthly CashFlowResult to an annual table.

    Treats any partial trailing year as an additional row, mirroring
    ``ProfitTester.profit_by_year``. The Year label for the partial row
    is the next integer (so a 241-month projection produces 20 full
    years plus a Year 21 row).
    """
    t = cf.projection_months or len(cf.net_cash_flow)
    if t == 0:
        return []

    def _sum_by_year(arr: np.ndarray) -> np.ndarray:
        n_full = t // 12
        remainder = t % 12
        full = (
            arr[: n_full * 12].reshape(-1, 12).sum(axis=1)
            if n_full > 0
            else np.array([], dtype=np.float64)
        )
        if remainder > 0:
            partial = arr[n_full * 12 :].sum()
            return np.append(full, partial)
        return full

    prem = _sum_by_year(cf.gross_premiums)
    claims = _sum_by_year(cf.death_claims)
    lapses = _sum_by_year(cf.lapse_surrenders)
    expenses = _sum_by_year(cf.expenses)
    reserve = _sum_by_year(cf.reserve_increase)
    ncf = _sum_by_year(cf.net_cash_flow)

    rows: list[dict[str, float]] = []
    for y in range(len(ncf)):
        rows.append(
            {
                "Year": y + 1,
                "Gross Premiums": float(prem[y]),
                "Death Claims": float(claims[y]),
                "Lapse Surrenders": float(lapses[y]),
                "Expenses": float(expenses[y]),
                "Reserve Increase": float(reserve[y]),
                "Net Cash Flow": float(ncf[y]),
            }
        )
    return rows


def _write_assumptions_sheet(wb: "Workbook", export: DealPricingExport) -> None:
    from openpyxl.styles import Font

    ws = wb.create_sheet(title="Assumptions")
    header_font = Font(bold=True)
    title_font = Font(bold=True, size=14)

    ws.cell(row=1, column=1, value="Deal & Assumption Metadata").font = title_font

    deal = export.deal_meta
    asm = export.assumptions_meta

    # Ordered list of (label, value, number_format). None format → default.
    rows: list[tuple[str, object, str | None]] = [
        ("Product Type", deal.product_type, None),
        ("Policies", deal.n_policies, "#,##0"),
        ("Face Amount", deal.face_amount, "$#,##0"),
        ("Treaty Type", deal.treaty_type if deal.treaty_type is not None else "None", None),
        (
            "Cession Percent",
            deal.cession_pct if deal.cession_pct is not None else "N/A",
            "0.00%" if deal.cession_pct is not None else None,
        ),
        ("Hurdle Rate", deal.hurdle_rate, "0.00%"),
        ("Discount Rate", deal.discount_rate, "0.00%"),
        ("Projection Years", deal.projection_years, "#,##0"),
        ("Valuation Date", deal.valuation_date.isoformat(), None),
        ("Mortality Source", asm.mortality_source, None),
        ("Mortality Multiplier", asm.mortality_multiplier, "0.000"),
        ("Lapse Description", asm.lapse_description, None),
        ("Assumption Set Version", asm.assumption_set_version, None),
    ]

    for i, (label, value, fmt) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = header_font
        cell = ws.cell(row=i, column=2, value=value)
        if fmt is not None:
            cell.number_format = fmt

    next_row = 3 + len(rows)
    if export.rated_block is not None and export.rated_block.n_rated > 0:
        _write_rated_block_panel(ws, export.rated_block, start_row=next_row + 1)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32


def _write_rated_block_panel(
    ws: "Worksheet",
    rated: RatedBlockExport,
    *,
    start_row: int,
) -> None:
    """Append a block-rating panel to the Assumptions sheet (ADR-068).

    Mirrors the labels and ordering of the CLI Rich
    ``_render_rated_block_table`` so committee reviewers see the same
    numbers across surfaces. ``start_row`` is the row of the section
    title; subsequent rows hold one labelled metric each.
    """
    from openpyxl.styles import Font

    section_font = Font(bold=True, italic=True)
    header_font = Font(bold=True)

    ws.cell(row=start_row, column=1, value="Rated Block").font = section_font

    panel_rows: list[tuple[str, object, str | None]] = [
        ("Policies Rated", rated.n_rated, "#,##0"),
        ("% Rated (by count)", rated.pct_rated_by_count, "0.0%"),
        ("% Rated (by face)", rated.pct_rated_by_face, "0.0%"),
        ("Face-weighted Avg Multiplier", rated.face_weighted_mean_multiplier, "0.000"),
        ("Max Multiplier", rated.max_multiplier, "0.00"),
        ("Max Flat Extra / $1,000", rated.max_flat_extra_per_1000, "$#,##0.00"),
    ]
    for offset, (label, value, fmt) in enumerate(panel_rows, start=1):
        row_idx = start_row + offset
        ws.cell(row=row_idx, column=1, value=label).font = header_font
        cell = ws.cell(row=row_idx, column=2, value=value)
        if fmt is not None:
            cell.number_format = fmt


def _write_sensitivity_sheet(wb: "Workbook", scenarios: list[ScenarioMetric]) -> None:
    from openpyxl.styles import Alignment, Font

    ws = wb.create_sheet(title="Sensitivity")
    header_font = Font(bold=True)
    centre = Alignment(horizontal="center")

    headers = ("Scenario", "PV Profits", "IRR", "Profit Margin")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = centre

    for i, s in enumerate(scenarios, start=2):
        ws.cell(row=i, column=1, value=s.name)
        pv_cell = ws.cell(row=i, column=2, value=float(s.pv_profits))
        pv_cell.number_format = "$#,##0"
        irr_cell = ws.cell(row=i, column=3, value=_fmt_rate(s.irr))
        if isinstance(irr_cell.value, float):
            irr_cell.number_format = "0.00%"
        margin_cell = ws.cell(row=i, column=4, value=_fmt_rate(s.profit_margin))
        if isinstance(margin_cell.value, float):
            margin_cell.number_format = "0.00%"

    ws.column_dimensions["A"].width = 28
    for col in "BCD":
        ws.column_dimensions[col].width = 16


def _write_yrt_rate_table_sheet(wb: "Workbook", table: "YRTRateTable") -> None:
    """Render the loaded YRT rate table on a single ``YRT Rate Table`` sheet.

    Layout: one block per (sex, smoker) cohort, stacked vertically with a
    blank row between blocks. Each block's header row labels the cohort
    and each rate row is ``[age, dur_1, dur_2, ..., dur_N, ultimate]``
    so the sheet is human-readable next to the source CSV (ADR-052).

    Filled cells (rows whose rate was forward/back-filled by
    ``YRTRateSchedule.generate_table`` rather than directly solved by
    brentq) are rendered in italic with a light-grey fill so reviewers
    can distinguish them from authoritative cells (ADR-054). When any
    cohort carries a ``solved_mask`` and at least one cell is filled, a
    note row is inserted below the title explaining the convention.
    CSV-loaded tables (mask ``None``) render exactly as before.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet(title="YRT Rate Table")
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    cohort_font = Font(bold=True, italic=True)
    centre = Alignment(horizontal="center")
    filled_font = Font(italic=True, color="666666")
    filled_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

    ws.cell(row=1, column=1, value=f"YRT Rate Table — {table.table_name}").font = title_font
    ws.cell(
        row=2,
        column=1,
        value=(
            f"Ages {table.min_age}-{table.max_age}, "
            f"select period {table.select_period_years} years. "
            f"Rates are annual $/$1,000 NAR."
        ),
    )

    has_any_filled = any(
        arr.solved_mask is not None and not arr.is_fully_solved for arr in table.arrays.values()
    )
    if has_any_filled:
        ws.cell(
            row=3,
            column=1,
            value=(
                "Italic / grey-filled cells were forward/back-filled from a "
                "solved row (age not directly solved; ADR-054)."
            ),
        ).font = Font(italic=True, color="666666")

    select_period = table.select_period_years
    headers = ["Age"] + [f"dur_{i}" for i in range(1, select_period + 1)] + ["ultimate"]

    row_idx = 5 if has_any_filled else 4
    for key in sorted(table.arrays.keys()):
        arr = table.arrays[key]
        # Cohort label.
        ws.cell(row=row_idx, column=1, value=f"Cohort: {key}").font = cohort_font
        row_idx += 1
        # Header row.
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = centre
        row_idx += 1
        # Data rows: one per attained age in the cohort.
        mask = arr.solved_mask
        for age_offset in range(arr.rates.shape[0]):
            age = arr.min_age + age_offset
            ws.cell(row=row_idx, column=1, value=age)
            for col_offset in range(arr.rates.shape[1]):
                rate_cell = ws.cell(
                    row=row_idx,
                    column=2 + col_offset,
                    value=float(arr.rates[age_offset, col_offset]),
                )
                rate_cell.number_format = "0.0000"
                is_solved = True if mask is None else bool(mask[age_offset, col_offset])
                if not is_solved:
                    rate_cell.font = filled_font
                    rate_cell.fill = filled_fill
            row_idx += 1
        # Blank row between cohorts.
        row_idx += 1

    ws.column_dimensions["A"].width = 12
    # `get_column_letter(N)` handles columns past Z (PR #39 P1 fix —
    # `chr(ord("B") + col_offset)` silently corrupted the column-width
    # map for `select_period >= 25`, which is in-range per
    # `YRTRateTable.select_period_years` (le=50).
    for col_offset in range(select_period + 1):
        ws.column_dimensions[get_column_letter(col_offset + 2)].width = 14


def write_yrt_rate_table_excel(table: "YRTRateTable", path: Path) -> None:
    """Write a standalone Excel workbook for a generated ``YRTRateTable``.

    Used by ``polaris rate-schedule --table`` to emit the solved tabular
    schedule as a deliverable workbook (ADR-053). Layout:

    * ``Summary`` sheet — table name, age range, select period, cohort
      list, and a count of solved cells.
    * ``YRT Rate Table`` sheet — one block per (sex, smoker) cohort
      stacked vertically (delegates to ``_write_yrt_rate_table_sheet``
      so the layout matches the deal-pricing workbook's appended sheet
      verbatim).

    Args:
        table: The populated ``YRTRateTable`` to render.
        path:  Output ``.xlsx`` file path.

    Raises:
        ImportError: If openpyxl is not installed (install via
            ``uv sync --extra tables``).
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise ImportError(
            "openpyxl required for Excel output. Install via: uv sync --extra tables"
        ) from exc

    wb = Workbook()
    if wb.active is not None:
        wb.remove(wb.active)

    # Summary sheet first so it opens by default.
    ws_summary = wb.create_sheet(title="Summary")
    ws_summary.cell(row=1, column=1, value=f"YRT Rate Table — {table.table_name}").font = Font(
        bold=True, size=14
    )
    ws_summary.cell(row=3, column=1, value="Generated by Polaris RE")
    ws_summary.cell(row=4, column=1, value=f"Age range: {table.min_age}-{table.max_age}")
    ws_summary.cell(row=5, column=1, value=f"Select period (years): {table.select_period_years}")
    ws_summary.cell(row=6, column=1, value=f"Cohorts: {len(table.arrays)}")
    n_cells = sum(int(arr.rates.size) for arr in table.arrays.values())
    ws_summary.cell(row=7, column=1, value=f"Total rate cells: {n_cells}")

    # Solved / filled disclosure (ADR-054). Only emitted when the table
    # was generated with per-cell solver provenance — CSV-loaded tables
    # leave ``solved_mask`` as None and continue to render only the
    # rows above.
    n_solved = 0
    n_filled = 0
    any_mask = False
    for arr in table.arrays.values():
        if arr.solved_mask is None:
            continue
        any_mask = True
        solved_count = int(arr.solved_mask.sum())
        n_solved += solved_count
        n_filled += int(arr.solved_mask.size) - solved_count
    if any_mask:
        ws_summary.cell(row=8, column=1, value=f"Solved cells: {n_solved}")
        ws_summary.cell(row=9, column=1, value=f"Filled cells: {n_filled}")
        ws_summary.cell(
            row=10,
            column=1,
            value=(
                "Filled cells were forward/back-filled from a solved row "
                "(ADR-054) — they are visually distinguished on the rate "
                "sheet by italic text and a light-grey fill."
            ),
        ).font = Font(italic=True, color="666666")

    ws_summary.column_dimensions["A"].width = 40

    _write_yrt_rate_table_sheet(wb, table)

    wb.save(path)
